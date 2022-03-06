import datetime
from datetime import date as Date
from datetime import time as Time
import io
from pyexpat import model
import random
from cv2 import split
import xlsxwriter as xlsxwriter
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User, Group
from django.http import FileResponse
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect
from django.template.defaulttags import register
from django.views.generic import UpdateView, ListView, DetailView, CreateView
from .backend.mishmar.Organizer import Organizer
from .forms import SettingsForm
from .models import ArmingRequest, IpBan, Post, ValidationLog
from .models import Settings as Settings
from .models import Event
from .models import Arming_Log
from .models import Gun
from .models import OrganizationShift
from .models import Organization1 as Organization
from .models import Shift as Shift
from users.models import UserSettings as USettings
import openpyxl
from django.utils import timezone
from openpyxl.utils import get_column_letter
from deep_translator import GoogleTranslator
import os
from django.views.generic.dates import DayArchiveView, MonthArchiveView
from django.utils import translation
from .decorators import user_staff_permission





default_language = os.environ.get("DEFAULT_LANGUAGE")

base_strings = {1: "הגשת משמרות", 2: "סידור", 3: "סידורים", 4: "הגדרות", 5: "ניהול נתונים", 6: "פרופיל", 7: "התנתק",
                8: "התחבר", 9: "הירשם"}

# initialize default shift numbers for suggestion organization
guards_num = {}
for x in range(14):
    guards_num[f"M{x}"] = 5
    guards_num[f"A{x}"] = 3
    guards_num[f"N{x}"] = 1
guards_num["A13"] = 0
guards_num["A12"] = 0
guards_num["M12"] = 0
guards_num["A6"] = 0
guards_num["M5"] = 0
guards_num["A5"] = 0
guards_num["N13"] = 2
guards_num["N12"] = 2
guards_num["M13"] = 2
guards_num["N6"] = 2
guards_num["N5"] = 2
guards_num["M6"] = 2


# Website settings view
@user_staff_permission
def settings_view(request):
    settings = Settings.objects.all().last()
    if request.method == 'POST':
        success, fail = ['שינויים נשמרו!', 'שינויים לא נשמרו!']
        settings_form = SettingsForm(request.POST, instance=settings)
        if settings_form.is_valid():
            messages.success(request, translate_text(success, request.user, "hebrew"))
            settings_form.save()
        else:
            messages.error(request, translate_text(fail, request.user, "hebrew"))
    else:
        settings_form = SettingsForm(instance=settings)
    context = {
        "settings_form": settings_form,
        "base": base_strings,
        "settings": settings,
    }
    return render(request, "mishmar/settings.html", context)

def start_app():
    if not Group.objects.filter(name='manager').exists():
        Group.objects.create(name='manager')

    if not Group.objects.filter(name='staff').exists():
        Group.objects.create(name='staff')

    if not User.objects.filter(username='metagber').exists():
        User.objects.create_user(username='metagber', password='mishmar123')

    if len(Settings.objects.all()) == 0:
        new_settings = Settings(submitting=True, pin_code=1234, officer="", city="", max_seq0=2, max_seq1=2, friday_morning=False, friday_noon=False)
        new_settings.save()

# Home view
def home(request):
    start_app()
    if request.user.is_authenticated:
        profile = USettings.objects.all().filter(user=request.user).last()
    else:
        profile = None
    posts = Post.objects.all()
    armingrequests = ArmingRequest.objects.all().filter(read=False)
    if request.user.is_authenticated:
        if request.user.groups.filter(name='manager').exists():
            num_requests = len(armingrequests)
            if num_requests > 0:
                messages.info(request, f'יש {num_requests} בקשות לשינוי ביומן חימוש')
        if request.user.groups.filter(name='staff').exists():
            usage = calculate_usage()
            if usage[0] >= 9000:
                messages.error(request, f'כמות נתונים גבוהה מאוד ({usage[0]}) אנא בצע גיבוי ומחיקת נתונים.')
            elif usage[0] >= 8000:
                messages.warning(request, f'כמות נתונים גבוהה ({usage[0]}) אנא בצע גיבוי ומחיקת נתונים.')
    context = {
        "posts": posts,
        "profile": profile,
    }
    return render(request, "mishmar/Home.html", context)

# Error 404 custom view
def error_404_view(request, exception):
    return render(request, 'mishmar/404.html')

# Error 500 custom view
def error_500_view(request, exception):
    return render(request, 'mishmar/500.html')

# Arming log chnage request to authorize change view
class ArmingRequestDetailView(UserPassesTestMixin,DetailView):
    model = ArmingRequest
    template_name = 'mishmar/arming_request_detail.html'
    context_object_name = 'armingrequest'

    def test_func(self):
        return self.request.user.groups.filter(name='manager').exists()

    def get_context_data(self, **kwargs):
        ctx = super(ArmingRequestDetailView, self).get_context_data(**kwargs)
        ctx['arming'] = self.get_object().log
        ctx['arming_data'] = self.get_object().log.data[str(self.get_object().input_num)]
        ctx = put_arming_context(ctx)
        return ctx
    
    def post(self, request, *args, **kwargs):
        if 'reject' in request.POST:
            armingrequest = ArmingRequest.objects.all().filter(id=self.get_object().id).first()
            armingrequest.read = True
            armingrequest.save()
            messages.success(request, "הבקשה נדחתה בהצלחה")
            return redirect('arming-requests-list')
        else:
            log = self.get_object().log
            armingrequest = ArmingRequest.objects.all().filter(id=self.get_object().id).first()
            armingrequest.read = True
            log_data = log.data[str(self.get_object().input_num)]
            log_data["shift_num"] = self.get_object().shift_num
            log_data["time_in"] = self.get_object().time_in
            log_data["time_out"] = self.get_object().time_out
            log_data["gun_id"] = int(self.get_object().gun.id)
            log_data["gun_case"] = self.get_object().gun_case
            log_data["mag_case"] = self.get_object().mag_case
            log_data["hand_cuffs"] = self.get_object().hand_cuffs
            log_data["num_mags"] = self.get_object().num_mags
            log_data["keys"] = self.get_object().keys
            log_data["radio"] = self.get_object().radio
            log_data["radio_kit"] = self.get_object().radio_kit
            valid_in = request.POST.get('sig-dataUrl')
            valid_out = request.POST.get('sig-dataUrl_out')
            if valid_in != "Empty":
                log_data["valid_in"] = valid_in
            if valid_out != "Empty":
                log_data["valid_out"] = valid_out
            if self.get_object().signature_in != None:
                log_data["signature_in"] = self.get_object().signature_in
            if self.get_object().signature_out != None:
                log_data["signature_out"] = self.get_object().signature_out
            log.save()
            armingrequest.save()
            messages.success(request, "הבקשה טופלה בהצלחה")
            return redirect('arming-requests-list')

# Arming log change request list view
class ArmingRequestListView(UserPassesTestMixin, ListView):
    model = ArmingRequest
    template_name = 'mishmar/arming_request_list.html'
    context_object_name = 'armingrequests'
    ordering = ["read", "-log__date"]
    paginate_by = 8

    def test_func(self):
        return self.request.user.groups.filter(name='manager').exists()

    def get_context_data(self, **kwargs):
        context = super(ArmingRequestListView, self).get_context_data(**kwargs)
        return context


# Arming log change request to send to manager view
class ArmingRequestView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = ArmingRequest
    template_name = "mishmar/change-request.html"
    fields = "__all__"

    def test_func(self):
        user_id = self.request.session["user_id"]
        if self.request.user.id == user_id:
            return True
        return False

    def get_context_data(self, **kwargs):
        ctx = super(ArmingRequestView, self).get_context_data(**kwargs)
        ctx = put_arming_context(ctx)
        log_id = self.request.session['log_id']
        log = Arming_Log.objects.filter(id=log_id).first()
        user_name = self.request.user.first_name + " " + self.request.user.last_name
        ctx["user_name"] = user_name
        session_keyes = ["gun_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out"]
        for key in session_keyes:
            ctx[key] = self.request.session[key]
        ctx["date"] = Date(self.request.session["year"], self.request.session["month"], self.request.session["day"])
        ctx["gun_s"] = Gun.objects.filter(id=self.request.session["gun_id"]).first()
        ctx["arming"] = log
        input_num = self.request.session["input_num"]
        ctx["arming_data"] = log.data[str(input_num)]
        return ctx
    
    def post(self, request, *args, **kwargs):
        session_keyes = ["gun_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out"]
        new_request = ArmingRequest()
        log = Arming_Log.objects.filter(id=request.session["log_id"]).first()
        new_request.log = log
        gun_id = int(request.session["gun_id"])
        gun  = Gun.objects.filter(id=gun_id).first()
        new_request.gun = gun
        new_request.shift_num = request.session["shift_num"]
        new_request.id_num = request.session["id_num"]
        new_request.time_in = request.session["time_in"]
        new_request.num_mags = request.session["num_mags"]
        new_request.hand_cuffs = request.session["hand_cuffs"]
        new_request.gun_case = request.session["gun_case"]
        new_request.mag_case = request.session["mag_case"]
        new_request.keys = request.session["keys"]
        new_request.radio = request.session["radio"]
        new_request.radio_kit = request.session["radio_kit"]
        new_request.read = False
        date1 = Date(request.session["year"], request.session["month"], request.session["day"])
        if request.session["time_out"] != "":
            new_request.time_out = request.session["time_out"]
        sig_in = request.POST.get('sig-dataUrl')
        sig_out = request.POST.get('sig-dataUrl_out')
        if sig_in != "Empty":
            new_request.signature_in = sig_in
        if sig_out != "Empty":
            new_request.signature_out = sig_out
        reason = request.POST.get('reason')
        if reason != "" and reason != None:
            new_request.reason = reason
            for key in session_keyes:
                del request.session[key]
            new_request.save()
            messages.success(request, "הבקשה הועברה בהצלחה")
            return redirect('armingday', year=int(date1.strftime("%Y")), month=date1.strftime("%b"), day=int(date1.strftime("%d")))
        else:
            messages.error(request, "אנא מלא סיבת שינוי")
            return HttpResponseRedirect(request.path_info)

# Arming log Day view
class ArmingDayView(LoginRequiredMixin, DayArchiveView):
    queryset = Arming_Log.objects.all()
    date_field = "date"
    allow_future = True
    allow_empty = True
    template_name = "mishmar/arming.html"

    def get_context_data(self, **kwargs):
        ctx = super(ArmingDayView, self).get_context_data(**kwargs)
        translation.activate('he')
        self.request.session[translation.LANGUAGE_SESSION_KEY] = 'he'
        ctx = put_arming_context(ctx)
        user_name = self.request.user.first_name + " " + self.request.user.last_name
        months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
        date1 = Date(self.kwargs['year'], months.index(getmonth(self.kwargs['month'].lower())) + 1, self.kwargs['day'])
        validation_log = ValidationLog.objects.all().filter(date=date1).first()
        ctx["user_name"] = user_name
        ctx["validation_log"] = validation_log
        armingrequests = ArmingRequest.objects.all().filter(read=False)
        ctx["num_requests"] = len(armingrequests)
        return ctx
    
    def post(self, request, *args, **kwargs):
        shift = 0
        if "add" in request.POST:
            request.session["gun_id"] = request.POST.get(f"guns")
            if request.user.username != "metagber": 
                request.session["name"] = request.user.first_name + " " + request.user.last_name
            else:
                request.session["name"] = request.POST.get(f"user_name")
            request.session["user_id"] = request.user.id
            session_keyes = ["id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit"]
            int_keyes = ["num_mags", "hand_cuffs", "gun_case", "mag_case"]
            bool_keyes = ["keys", "radio", "radio_kit"]
            for key in session_keyes:
                if key in int_keyes:
                    request.session[key] = int(request.POST[f"{key}"])
                elif key in bool_keyes:
                    request.session[key] = checkbox(request.POST.get(f"{key}", None))
                else:
                    request.session[key] = request.POST[f"{key}"]
            request.session["shift_num"] = int(request.POST.get("shifts"))
            time_out = request.POST.get("time_out")
            if request.user.groups.filter(name="manager").exists():
                request.session["reqtype"] = "manager"
            else:
                request.session["reqtype"] = "add"
            if time_out != "":
                request.session["time_out"] = time_out
            else:
                request.session["time_out"] = ""
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            request.session["year"] = self.kwargs['year']
            request.session["month"] = months.index(getmonth(self.kwargs['month'].lower())) + 1
            request.session["day"] = self.kwargs['day']
            date1 = Date(self.kwargs['year'], months.index(getmonth(self.kwargs['month'].lower())) + 1, self.kwargs['day'])
            logs = Arming_Log.objects.all().filter(date=date1)
            if len(logs) > 0:
                request.session["log_id"] = logs.first().id
            else:
                request.session["log_id"] = ""
            messages.info(request, "הנתונים הועברו בהצלחה, יש לחתום כדי לשמור")
            return redirect("arming-new")
        elif "request" in request.POST:
            input_num = request.POST.get("request")
            request.session["input_num"] = input_num
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            date1 = Date(year=self.kwargs['year'], month=months.index(getmonth(self.kwargs['month'].lower())) + 1, day=self.kwargs['day'])
            log = Arming_Log.objects.get(date=date1)
            log_data = log.data[str(input_num)]
            request.session["gun_id"] = request.POST.get(f"guns{input_num}")
            request.session["shift_num"] = int(request.POST.get(f"shifts{input_num}"))
            request.session["user_id"] = request.user.id
            session_keyes = ["id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit"]
            int_keyes = ["num_mags", "hand_cuffs", "gun_case", "mag_case"]
            bool_keyes = ["keys", "radio", "radio_kit"]
            for key in session_keyes:
                if key in int_keyes:
                    request.session[key] = int(request.POST[f"{key}{input_num}"])
                elif key in bool_keyes:
                    request.session[key] = checkbox(request.POST.get(f"{key}{input_num}", None))
                else:
                    request.session[key] = request.POST[f"{key}{input_num}"]
            time_out = request.POST.get(f"time_out{input_num}")
            request.session["log_id"] = log.id
            if request.user.username != "metagber":
                name = request.user.first_name + " " + request.user.last_name
            else:
                name = request.POST.get(f"user_name{input_num}")
            request.session["name"] = log_data["name"]
            if time_out != "":
                request.session["time_out"] = time_out
            else:
                request.session["time_out"] = ""
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            request.session["year"] = self.kwargs['year']
            request.session["month"] = months.index(getmonth(self.kwargs['month'].lower())) + 1
            request.session["day"] = self.kwargs['day']
            messages.info(request, "אנא עבור על בקשתך עם הסבר ואשר")
            return redirect("arming-changerequest")
        elif "change" in request.POST:
            input_num = request.POST.get("change")
            request.session["input_num"] = input_num
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            date1 = Date(year=self.kwargs['year'], month=months.index(getmonth(self.kwargs['month'].lower())) + 1, day=self.kwargs['day'])
            log = Arming_Log.objects.get(date=date1)
            log_data = log.data[str(input_num)]
            request.session["gun_id"] = request.POST.get(f"guns{input_num}")
            request.session["shift_num"] = int(request.POST.get(f"shifts{input_num}"))
            request.session["user_id"] = request.user.id
            session_keyes = ["id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit"]
            int_keyes = ["num_mags", "hand_cuffs", "gun_case", "mag_case"]
            bool_keyes = ["keys", "radio", "radio_kit"]
            for key in session_keyes:
                if key in int_keyes:
                    request.session[key] = int(request.POST[f"{key}{input_num}"])
                elif key in bool_keyes:
                    request.session[key] = checkbox(request.POST.get(f"{key}{input_num}", None))
                else:
                    request.session[key] = request.POST[f"{key}{input_num}"]
            time_out = request.POST.get(f"time_out{input_num}")
            request.session["log_id"] = log.id
            request.session["input_num"] = request.POST.get("change")
            if request.user.username != "metagber":
                name = request.user.first_name + " " + request.user.last_name
            else:
                name = request.POST.get(f"user_name{input_num}")
            request.session["name"] = log_data["name"]
            if name == log_data["name"] and request.user.groups.filter(name="manager").exists():
                request.session["reqtype"] = "change manager"
            elif name == log_data["name"]:
                request.session["reqtype"] = "change"
            else:
                request.session["reqtype"] = "validation"
            if time_out != "":
                request.session["time_out"] = time_out
            else:
                request.session["time_out"] = ""
            messages.info(request, " הנתונים הועברו בהצלחה כדי לשמור יש לחתום")
            return redirect("signature", log.id)
        elif "month_log" in request.POST:
            return redirect("armingmonth", year=self.kwargs['year'], month=self.kwargs['month'])
        elif "month_log_manager" in request.POST:
            return redirect("armingmonth-all", year=self.kwargs['year'], month=self.kwargs['month'])
        elif "change_requests" in request.POST:
            return redirect("arming-requests-list")
        elif 'validationlog-month' in request.POST:
            return redirect("validation-month", year=self.kwargs['year'], month=self.kwargs['month'])
        elif "shift1" in request.POST:
            shift = 1
        elif "Shift" in request.POST:
            shift = 2
        elif "shift3" in request.POST:
            shift = 3
        elif "goto" in request.POST:
            date = request.POST.get("goto_date")
            date1 = Date(int(date[0:4]), int(date[5:7]), int(date[8:10]))
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            return redirect('armingday', year=date1.year, month=getmonth(date1.strftime("%b")), day=date1.day)
        if shift != 0:
            manager = request.POST.get(f"manager{shift}")
            if manager == "":
                messages.error(request, translate_text("נא למלא את שם האחמ\"ש", request.user, "hebrew"))
                return HttpResponseRedirect(request.path_info)
            val_logs = ValidationLog.objects.all()
            months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            request.session["year"] = self.kwargs['year']
            request.session["month"] = months.index(getmonth(self.kwargs['month'].lower())) + 1
            request.session["day"] = self.kwargs['day']
            date1 = Date(self.kwargs['year'], months.index(getmonth(self.kwargs['month'].lower())) + 1, self.kwargs['day'])
            log = val_logs.filter(date=date1)
            if len(log) == 0:
                request.session["reqtype"] = "add"
            else:
                request.session["reqtype"] = "change"
            session_keyes = ["gun_safe", "gun_shift", "time", "manager"]
            for key in session_keyes:
                request.session[key] = request.POST[f"{key}{shift}"]
            request.session["shift"] = shift
            messages.info(request, "הנתונים הועברו בהצלחה, יש לחתום בכדי לשמור")
            return redirect("validation-signature")

# Arming log update view
class ArmingLogUpdate(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Arming_Log
    template_name = "mishmar/signature_page.html"
    fields = []
    context_object_name = 'arming'

    def test_func(self):
        user_id = self.request.session["user_id"]
        if self.request.user.groups.filter(name="manager").exists():
            return True
        data = self.get_object().data[str(self.request.session[f"input_num"])]
        if data["user_id"] == user_id:
            if data["shift_num"] != 3:
                if datetime.datetime.now().date() == self.get_object().date:
                    return True
            else:
                if datetime.datetime.now().date() == self.get_object().date or datetime.datetime.now().date() == self.get_object().date + datetime.timedelta(days=1):
                    return True
        return False

    def get_context_data(self, **kwargs):
        ctx = super(ArmingLogUpdate, self).get_context_data(**kwargs)
        guns = Gun.objects.all()
        ctx = put_arming_context(ctx)
        user_name = self.request.session["name"]
        ctx["user_name"] = user_name
        session_keyes = ["gun_id","name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out", "reqtype"]
        for key in session_keyes:
            ctx[key] = self.request.session[key]
        input_num = self.request.session["input_num"]
        arming_data = self.get_object().data[str(input_num)]
        ctx["arming_data"] = arming_data
        gun_id = self.request.session["gun_id"]
        short_name = guns.filter(id=gun_id).first().short_name
        ctx["short_name"] = short_name
        return ctx
    
    def post(self, request, *args, **kwargs):
        reqtype = request.session["reqtype"]
        session_keyes = ["gun_id", "user_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out", "reqtype"]
        int_keys = ["gun_id", "user_id", "shift_num", "num_mags", "hand_cuffs", "gun_case", "mag_case"]
        log = Arming_Log.objects.filter(id=self.get_object().id).first()
        input_num = request.session["input_num"]
        for i in range(len(session_keyes) - 1):
            if session_keyes[i] in int_keys:
                log.data[str(input_num)][f'{session_keyes[i]}'] = int(request.session[session_keyes[i]])
            else:
                log.data[str(input_num)][f'{session_keyes[i]}'] = request.session[session_keyes[i]]
        sig_in = request.POST.get('sig-dataUrl')
        sig_out = request.POST.get('sig-dataUrl_out')
        if reqtype == "change manager":
            valid_in = request.POST.get('sig-dataUrl_valid')
            valid_out = request.POST.get('sig-dataUrl_out_valid')
            if sig_in != "Empty":
                log.data[str(input_num)][f'signature_in'] = sig_in
            if sig_out != "Empty":
                log.data[str(input_num)][f'signature_out'] = sig_out
            if valid_in != "Empty":
                log.data[str(input_num)][f'valid_in'] = valid_in
            if valid_out != "Empty":
                log.data[str(input_num)][f'valid_out'] = valid_out
        elif reqtype == "change":
            if sig_in != "Empty":
                log.data[str(input_num)][f'signature_in'] = sig_in
            if sig_out != "Empty":
                log.data[str(input_num)][f'signature_out'] = sig_out
        else:
            if sig_in != "Empty":
                log.data[str(input_num)][f'valid_in'] = sig_in
            if sig_out != "Empty":
                log.data[str(input_num)][f'valid_out'] = sig_out
        if sig_in == "Empty" and sig_out == "Empty" and reqtype != "change manager":
            messages.warning(request, "אנא הכנס את החתימה שלך")
            return HttpResponseRedirect(request.path_info)
        elif reqtype == "change manager" and sig_in == "Empty" and sig_out == "Empty" and valid_in == "Empty" and valid_out == "Empty":
            messages.warning(request, "אנא הכנס את החתימה שלך")
            return HttpResponseRedirect(request.path_info)
        else:
            for key in session_keyes:
                del request.session[key]
            log.save()
            messages.success(request, "הנתונים נשמרו בהצלחה")
            return redirect('armingday', year=int(self.get_object().date.strftime("%Y")), month=self.get_object().date.strftime("%b"), day=int(self.get_object().date.strftime("%d")))


# Arming log create view
class ArmingCreateView(LoginRequiredMixin, CreateView):
    model = Arming_Log
    template_name = "mishmar/signature_create.html"
    fields = "__all__"

    def get_context_data(self, **kwargs):
        ctx = super(ArmingCreateView, self).get_context_data(**kwargs)
        ctx = put_arming_context(ctx)
        user_name = self.request.user.first_name + " " + self.request.user.last_name
        ctx["user_name"] = user_name
        session_keyes = ["gun_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out", "reqtype"]
        for key in session_keyes:
            ctx[key] = self.request.session[key]
        ctx["date"] = Date(self.request.session["year"], self.request.session["month"], self.request.session["day"])
        ctx["gun_s"] = Gun.objects.filter(id=self.request.session["gun_id"]).first()
        return ctx
    
    def post(self, request, *args, **kwargs):
        session_keyes = ["gun_id","user_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out", "reqtype"]
        reqtype = request.session["reqtype"]
        date1 = Date(request.session["year"], request.session["month"], request.session["day"])
        log_id = request.session["log_id"]
        input_num = 1
        if log_id == "":
            log  = Arming_Log(date=date1)
            log.data = {}
            log.data["input_num"] = 1
            input_num = 1
            log.data[str(input_num)] = {}
            log.data[str(input_num)][f"signature_in"] = ""
            log.data[str(input_num)][f"signature_out="] = ""
            log.data[str(input_num)][f"valid_in"] = ""
            log.data[str(input_num)][f"valid_out"] = ""
        else:
            log = Arming_Log.objects.all().filter(id=log_id).first()
            input_num = int(log.data["input_num"]) + 1
            log.data["input_num"] = input_num
            log.data[str(input_num)] = {}
        for i in range(len(session_keyes) - 1):
            if f'{session_keyes[i]}' == "gun_id":
                log.data[str(input_num)][f'{session_keyes[i]}'] = int(request.session[session_keyes[i]])
            else:
                log.data[str(input_num)][f'{session_keyes[i]}'] = request.session[session_keyes[i]]
        sig_in = request.POST.get('sig-dataUrl')
        sig_out = request.POST.get('sig-dataUrl_out')
        signature = False
        if sig_in != "Empty":
            log.data[str(input_num)][f"signature_in"] = sig_in
            signature = True
        if sig_out != "Empty":
            log.data[str(input_num)][f"signature_out"] = sig_out
            signature = True
        if reqtype == "manager":
            valid_in = request.POST.get('sig-dataUrl_valid')
            valid_out = request.POST.get('sig-dataUrl_out_valid')
            if valid_in != "Empty":
                log.data[str(input_num)][f"valid_in"] = request.POST.get('sig-dataUrl_valid')
            if valid_out != "Empty":
                log.data[str(input_num)][f"valid_out"] = request.POST.get('sig-dataUrl_out_valid')
        if signature:
            for key in session_keyes:
                del request.session[key]
            log.save()
            return redirect('armingday', year=int(date1.strftime("%Y")), month=date1.strftime("%b"), day=int(date1.strftime("%d")))
        else:
            messages.warning(request, "אנא הכנס את החתימה שלך")
            return HttpResponseRedirect(request.path_info)

# View to create or edit Validation log 
@login_required
def Validation_Log_Signature(request):
    context = {}
    session_keyes = ["gun_safe", "gun_shift", "time", "manager", "reqtype", "shift", "year", "month", "day"]
    for key in session_keyes:
        context[key] = request.session[key]
    date1 = Date(int(context["year"]), int(context["month"]), int(context["day"]))
    context["date"] = date1
    log = ValidationLog.objects.filter(date=date1).first()
    context["denied"] = False
    if log != None:
        context["denied"] = validation_log_check(log, context["shift"])
    if request.method == "POST":
        sig = request.POST.get('sig-dataUrl')
        if sig == "Empty":
            messages.warning(request, "אנא הכנס את החתימה שלך")
            return HttpResponseRedirect(request.path_info)
        else:
            shift = context["shift"]
            if shift == 1:
                shift = "m"
            elif shift == 2:
                shift = "a"
            else:
                shift = "n"
            if context["reqtype"] == "add":
                log = ValidationLog()
                log.date = date1
            else:
                log = ValidationLog.objects.filter(date=date1).first()
            log.__setattr__(f"num_guns_safe_{shift}", context["gun_safe"])
            log.__setattr__(f"num_guns_shift_{shift}", context["gun_shift"])
            log.__setattr__(f"time_checked_{shift}", context["time"])
            log.__setattr__(f"name_checked_{shift}", context["manager"])
            log.__setattr__(f"sig_{shift}", sig)
            for key in session_keyes:
                del request.session[key]
            log.save()
            messages.success(request, "החתימה נשמרה בהצלחה")
            return redirect('armingday', year=int(date1.strftime("%Y")), month=date1.strftime("%b"), day=int(date1.strftime("%d")))
    return render(request, "mishmar/validation_signature.html", context)

# Validation log month view
class ValidationLogMonthView(LoginRequiredMixin, UserPassesTestMixin, MonthArchiveView):
    queryset = ValidationLog.objects.all()
    date_field = "date"
    allow_future = True
    allow_empty = True
    template_name = "mishmar/validation-month.html"
    ordering = ["date"]

    def test_func(self):
        return isStaff(self.request.user)

# Arming log month view
class ArmingPersonalMonthView(LoginRequiredMixin, MonthArchiveView):
    queryset = Arming_Log.objects.all()
    date_field = "date"
    allow_future = True
    allow_empty = True
    template_name = "mishmar/arming-month.html"
    ordering = ["date"]

    def get_context_data(self, **kwargs):
        ctx = super(ArmingPersonalMonthView, self).get_context_data(**kwargs)
        ctx = put_arming_context(ctx)
        user_name = self.request.user.first_name + " " + self.request.user.last_name
        ctx["user_name"] = user_name
        ctx["reqtype"] = ""
        return ctx
    
    def post(self, request, *args, **kwargs):
        pass

# Arming log month view
class ArmingAllMonthView(LoginRequiredMixin, MonthArchiveView):
    queryset = Arming_Log.objects.all()
    date_field = "date"
    allow_future = True
    allow_empty = True
    template_name = "mishmar/arming-month.html"
    ordering = ["date"]

    def get_context_data(self, **kwargs):
        ctx = super(ArmingAllMonthView, self).get_context_data(**kwargs)
        ctx = put_arming_context(ctx)
        user_name = self.request.user.first_name + " " + self.request.user.last_name
        ctx["user_name"] = user_name
        ctx["reqtype"] = 'manager'
        return ctx
    
    def post(self, request, *args, **kwargs):
        pass


# Guns view
class GunListView(LoginRequiredMixin,UserPassesTestMixin, ListView):
    model = Gun
    template_name = "mishmar/gun-list.html"
    context_object_name = "guns"

    def test_func(self):
        return isStaff(self.request.user)

    def get_context_data(self, **kwargs):
        ctx = super(GunListView, self).get_context_data(**kwargs)
        return ctx
    
    def post(self, request, *args, **kwargs):
        if 'new' in request.POST:
            new_gun = Gun()
            new_gun.full_name = self.request.POST.get('long')
            new_gun.short_name = self.request.POST.get('short')
            new_gun.save()
            messages.success(request, "אקדח נוצר")
        elif 'delete' in request.POST:
            id = request.POST.get("delete")
            gun = Gun.objects.all().filter(id=id).first()
            gun.delete()
            messages.success(request, "אקדח נמחק")
        else:
            id = request.POST.get("change")
            gun = Gun.objects.all().filter(id=id).first()
            gun.full_name = self.request.POST.get(f'long{id}')
            gun.short_name = self.request.POST.get(f'short{id}')
            gun.save()
            messages.success(request, "אקדח שונה")
        return HttpResponseRedirect(request.path_info)


# View to serv shifts to create or edit ShiftWeek and Shift
@login_required
def shift_view(request):
    organization = Organization.objects.order_by('-date')[0]
    notes_text = ""
    empty = False
    settings = Settings.objects.last()
    submitting = settings.submitting
    user_settings = USettings.objects.all().filter(user=request.user).first()
    days = []
    for x in range(organization.num_weeks * 7):
        days.append(Organization.objects.order_by('-date')[0].date + datetime.timedelta(days=x))
    events = Event.objects.all()
    for x in range(organization.num_weeks * 7):
        if len(events.filter(date2=days[x])) > 0:
            for ev in events.filter(date2=days[x]):
                if user_settings.nickname == ev.nickname:
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}. אם יש שינוי להודיע!'
                    message = translate_text(message, request.user, "hebrew")
                    messages.info(request, message)
                elif ev.nickname == 'כולם':
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}'
                    message = translate_text(message, request.user, "hebrew")
                    messages.info(request, message)
                elif ev.nickname == 'מנהלים' and isStaff(request.user):
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}'
                    message = translate_text(message, request.user, "hebrew")
                    messages.info(request, message)
                elif ev.nickname == 'אחמשים' and is_manager(request.user):
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}'
                    message = translate_text(message, request.user, "hebrew")
                    messages.info(request, message)
    if request.method == 'POST':
        if not already_submitted(request.user):
            shift = Shift()
            shift_data_temp = {}
            for i in range(organization.num_weeks):
                shift_data_temp[str(i)] = {}
                for j in range(1, 8):
                    shift_data_temp[str(i)][f'M{j}'] = False
                    shift_data_temp[str(i)][f'P{j}'] = True
                    shift_data_temp[str(i)][f'A{j}'] = False
                    shift_data_temp[str(i)][f'N{j}'] = False
                    shift_data_temp[str(i)][f'R{j}'] = False
                    shift_data_temp[str(i)][f'notes{j}'] = ""
            shift.weeks_data = shift_data_temp
        else:
            org = Organization.objects.order_by('-date')[0]
            shifts = Shift.objects.filter(organization=org)
            shift = shifts.filter(user=request.user).first()
            notes_text = str(shift.notes)
        shift.user = request.user
        shift.organization = Organization.objects.order_by('-date').first()
        notes_area = request.POST.get("notesArea")
        shift.notes = notes_area
        shift.seq_night = request.POST.get(f"seq_night", 0)
        shift.seq_noon = request.POST.get(f"seq_noon", 0)
        shift_data_temp = {}
        for i in range(organization.num_weeks):
            shift_data_temp[str(i)] = {}
            for j in range(1, 8):
                shift_data_temp[str(i)][f'M{j}'] = False
                shift_data_temp[str(i)][f'P{j}'] = True
                shift_data_temp[str(i)][f'A{j}'] = False
                shift_data_temp[str(i)][f'N{j}'] = False
                shift_data_temp[str(i)][f'R{j}'] = False
                shift_data_temp[str(i)][f'notes{j}'] = ""
        for i in range(organization.num_weeks):
            for j in range(1, 8):
                shift_data_temp[str(i)][f"M{j}"] = checkbox(request.POST.get(f"M{j}_{i}", False))
                shift_data_temp[str(i)][f"P{j}"] = checkbox(request.POST.get(f"P{j}_{i}", False))
                shift_data_temp[str(i)][f"A{j}"] = checkbox(request.POST.get(f"A{j}_{i}", False))
                shift_data_temp[str(i)][f"N{j}"] = checkbox(request.POST.get(f"N{j}_{i}", False))
                shift_data_temp[str(i)][f"R{j}"] = checkbox(request.POST.get(f"R{j}_{i}", False))
                shift_data_temp[str(i)][f"notes{j}"] = request.POST.get(f"notes{j}_{i}", False)
        if not already_submitted(request.user):
            messages.success(request, translate_text(f'משמרות הוגשו בהצלחה!', request.user, "hebrew"))
        else:
            messages.success(request, translate_text(f'משמרות עודכנו בהצלחה!', request.user, "hebrew"))
        shift.weeks_data = shift_data_temp
        shift.save()
        shift_data = shift.weeks_data
        return redirect("Home")
    else:
        if not submitting:
            shifts = Shift.objects.order_by('-organization__date')
            if len(shifts.filter(user=request.user, organization=organization).order_by('-organization__date')) > 0:
                shift = shifts.filter(user=request.user).order_by('-organization__date').first()
                shift_data = shift.weeks_data
                notes_text = str(shift.notes)
            else:
                empty = True
        elif not already_submitted(request.user):
            shift_data = {}
            shift = None
            for i in range(organization.num_weeks):
                shift_data[str(i)] = {}
                for j in range(1, 8):
                    shift_data[str(i)][f"M{j}"] = False
                    shift_data[str(i)][f"P{j}"] = True
                    shift_data[str(i)][f"A{j}"] = False
                    shift_data[str(i)][f"N{j}"] = False
                    shift_data[str(i)][f"R{j}"] = False
                    shift_data[str(i)][f"notes{j}"] = ""
            notes_text = ""
        else:
            shifts = Shift.objects.filter(organization=organization)
            shift = shifts.filter(user=request.user).first()
            shift_data = shift.weeks_data
            notes_text = str(shift.notes)
    if not empty:
        context = {
            "shift": shift,
            "shift_data": shift_data,
            "days": days,
            "submitting": submitting,
            "notes_text": notes_text,
            "empty": empty,
            "manager": False,
        }
    else:
        context = {
            "empty": empty,
            "manager": False,
        }
    return render(request, "mishmar/shifts.html", context)


# Organizations list view by date
class ServedSumListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Organization
    template_name = "mishmar/Served_sum_list.html"
    context_object_name = "organizations"
    ordering = ["-date"]
    paginate_by = 5

    def test_func(self):
        return isStaff(self.request.user)


# View to staff member to edit shifts served 
@user_staff_permission
def shift_update_view(request, pk=None):
    shift = Shift.objects.all().filter(id=pk).first()
    shift_data = shift.weeks_data
    organization = shift.organization
    user = shift.user
    notes_text = str(shift.notes)
    user_settings = USettings.objects.all().filter(user=user).first()
    days = []
    for x in range(organization.num_weeks * 7):
        days.append(organization.date + datetime.timedelta(days=x))
    events = Event.objects.all()
    for x in range(organization.num_weeks * 7):
        if len(events.filter(date2=days[x])) > 0:
            for ev in events.filter(date2=days[x]):
                if user_settings.nickname == ev.nickname:
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}. אם יש שינוי להודיע!'
                    message = translate_text(message, user, "hebrew")
                    messages.info(request, message)
                elif ev.nickname == 'כולם':
                    message = f'לא לשכוח בתאריך {ev.date2} יש {ev.description}'
                    message = translate_text(message, user, "hebrew")
                    messages.info(request, message)
    if request.method == 'POST':
        notes_area = request.POST.get("notesArea")
        shift.notes = notes_area
        notes_text = str(shift.notes)
        shift.seq_night = request.POST.get(f"seq_night", 0)
        shift.seq_noon = request.POST.get(f"seq_noon", 0)
        shift_data_temp = {}
        for i in range(organization.num_weeks):
            shift_data_temp[str(i)] = {}
            for j in range(1, 8):
                shift_data_temp[str(i)][f'M{j}'] = False
                shift_data_temp[str(i)][f'P{j}'] = True
                shift_data_temp[str(i)][f'A{j}'] = False
                shift_data_temp[str(i)][f'N{j}'] = False
                shift_data_temp[str(i)][f'R{j}'] = False
                shift_data_temp[str(i)][f'notes{j}'] = ""
        for i in range(organization.num_weeks):
            for j in range(1, 8):
                shift_data_temp[str(i)][f"M{j}"] = checkbox(request.POST.get(f"M{j}_{i}", False))
                shift_data_temp[str(i)][f"P{j}"] = checkbox(request.POST.get(f"P{j}_{i}", False))
                shift_data_temp[str(i)][f"A{j}"] = checkbox(request.POST.get(f"A{j}_{i}", False))
                shift_data_temp[str(i)][f"N{j}"] = checkbox(request.POST.get(f"N{j}_{i}", False))
                shift_data_temp[str(i)][f"R{j}"] = checkbox(request.POST.get(f"R{j}_{i}", False))
                shift_data_temp[str(i)][f"notes{j}"] = request.POST.get(f"notes{j}_{i}", False)
        shift.weeks_data = shift_data_temp
        shift.save()
        messages.success(request, translate_text(f'משמרות עודכנו בהצלחה!', user, "hebrew"))
        return redirect("Served-sum")
    context = {
            "shift": shift,
            "days": days,
            "submitting": True,
            "notes_text": notes_text,
            "shift_data": shift_data,
            "empty": False,
            "manager": True,
            "userview": USettings.objects.all().filter(user=user).first().nickname,
    }
    return render(request, "mishmar/shifts.html", context)


# View to show staff memeber ready form of the organization
class OrganizationDetailView(LoginRequiredMixin, DetailView, UserPassesTestMixin):
    model = Organization
    template_name = "mishmar/organization-detail.html"

    def test_func(self):
        return isStaff(self.request.user)
    
    def get_context_data(self, **kwargs):
        ctx = super(OrganizationDetailView, self).get_context_data(**kwargs)
        shifts = OrganizationShift.objects.all().order_by('shift_num', 'index')
        weeks = []
        for i in range(self.get_object().num_weeks):
            weeks.append(self.get_object().weeks_data[str(i)])
        ctx["weeks"] = weeks
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ctx["days"] = days
        ctx["shifts"] = shifts
        return ctx


# View to create an organization with date and num weeks
class OrganizationCreateView(LoginRequiredMixin, CreateView, UserPassesTestMixin):
    model = Organization
    template_name = "mishmar/organization-new.html"
    fields = ('date', 'num_weeks')

    def get_context_data(self, **kwargs):
        ctx = super(OrganizationCreateView, self).get_context_data(**kwargs)
        organizations = Organization.objects.all().order_by('-date')
        if organizations.count() > 0:
            ctx["date1"] = organizations.first().date + datetime.timedelta(days=7 * organizations.first().num_weeks)
        else:
            ctx["date1"] = timezone.now()
        return ctx

    def test_func(self):
        return isStaff(self.request.user)

    def form_valid(self, form):
        if Organization.objects.all().filter(date=self.request.POST.get("date")).exists():
            messages.error(self.request, translate_text(f'סידור בתאריך זה כבר קיים', self.request.user, "hebrew"))
            return redirect("organization-new")
        form.instance.date = self.request.POST.get("date")
        return super().form_valid(form)


# View to show staff memeber a table of shifts users got based on the organization
class ShifttableView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Organization
    template_name = "mishmar/shift-table-view.html"

    def get_context_data(self, **kwargs):
        ctx = super(ShifttableView, self).get_context_data(**kwargs)
        weeks = []
        for i in range(self.get_object().num_weeks):
            weeks.append(self.get_object().weeks_data[str(i)])
        users = set()
        shift_keys = OrganizationShift.objects.all().exclude(shift_num=4)
        for i in range(self.get_object().num_weeks):
            for key in shift_keys:
                for day in range(1, 8):
                    temp_shift = weeks[i][f'{day}@{key.id}'].replace("\r", "\n")
                    split = temp_shift.split("\n")
                    for s in split:
                        s = s.replace(" ", "")
                        if s != "":
                            users.add(s)
        if " " in users:
            users.remove(" ")
        table_content = {}
        sum_content = {}
        for i in range(self.get_object().num_weeks):
            sum_content[f'morning{i + 1}'] = 0
            sum_content[f'noon{i + 1}'] = 0
        sum_content["night"] = 0
        sum_content["end"] = 0
        for user in users:
            table_content[user] = {}
            for i in range(self.get_object().num_weeks):
                table_content[user][f'morning{i + 1}'] = 0
                table_content[user][f'noon{i + 1}'] = 0
            table_content[user]["night"] = 0
            table_content[user]["end"] = 0
        shift_keys = OrganizationShift.objects.all().order_by('shift_num', 'index')
        morning_shifts = shift_keys.filter(shift_num=1)
        noon_shifts = shift_keys.filter(shift_num=2)
        night_shifts = shift_keys.filter(shift_num=3)
        for j in range(self.get_object().num_weeks):
            for i in range(1, 8):
                for shift in morning_shifts:
                    split = weeks[j][f'{i}@{shift.id}'].replace("\r", "\n").split("\n")
                    for s in split:
                        s = s.replace(" ", "")
                        if s != "":
                            if i != 7:
                                table_content[s][f"morning{j + 1}"] += 1
                                sum_content[f"morning{j + 1}"] += 1
                            else:
                                table_content[s]["end"] += 1
                                sum_content["end"] += 1
                for shift in noon_shifts:
                    split = weeks[j][f'{i}@{shift.id}'].replace("\r", "\n").split("\n")
                    for s in split:
                        s = s.replace(" ", "")
                        if s != "":
                            if i < 6:
                                table_content[s][f"noon{j + 1}"] += 1
                                sum_content[f"noon{j + 1}"] += 1
                            else:
                                table_content[s]["end"] += 1
                                sum_content["end"] += 1
                for shift in night_shifts:
                    split = weeks[j][f'{i}@{shift.id}'].replace("\r", "\n").split("\n")
                    for s in split:
                        s = s.replace(" ", "")
                        if s != "":
                            if i < 6:
                                table_content[s]["night"] += 1
                                sum_content["night"] += 1
                            else:
                                table_content[s]["end"] += 1
                                sum_content["end"] += 1
        ctx["table"] = table_content
        ctx["sum"] = sum_content
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ctx["days"] = days
        ctx["num_weeks"] = self.get_object().num_weeks + 1
        return ctx

    def test_func(self):
        return isStaff(self.request.user)


# View to show staff member shift reinforcements served
class ServedSumReinforcementsDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Organization
    template_name = "mishmar/served_sum_reinforcements.html"

    def get_data2(self, calculated: bool):
        ctx = {}
        served = {}
        notes = []
        shift_keys = OrganizationShift.objects.all().order_by('shift_num', 'index')
        morning_shifts = shift_keys.filter(shift_num=1)
        noon_shifts = shift_keys.filter(shift_num=2)
        night_shifts = shift_keys.filter(shift_num=3)
        if calculated:
            organization = self.get_object()
            organization_input = get_input(organization)
            input_days = {}
        for i in range(1, self.get_object().num_weeks * 7 + 1):
            served["day" + str(i)] = ""
            if calculated:
                day = "day" + str(i)
                input_days[day + "M"] = []
                input_days[day + "A"] = []
                input_days[day + "N"] = []
                for key in morning_shifts:
                    input_days[day + "M"] += organization_input[f'{i}@{key.id}'].split("\n")
                for key in noon_shifts:
                    input_days[day + "A"] += organization_input[f'{i}@{key.id}'].split("\n")
                for key in night_shifts:
                    input_days[day + "N"] += organization_input[f'{i}@{key.id}'].split("\n")
        if calculated:
            for key in input_days:
                for i in range(len(input_days[key])):
                    input_days[key][i] = input_days[key][i].replace(" ", "")
                    input_days[key][i] = input_days[key][i].replace("\n", "")
                    input_days[key][i] = input_days[key][i].replace("\r", "")
        main_shifts_served = Shift.objects.all().filter(organization=self.get_object())
        weeks_notes = []
        for i in range(self.get_object().num_weeks):
            weeks_notes.append("")
        notes_general = ""
        users = {}
        user_notes_added = []
        for shift in main_shifts_served:
            user = shift.user
            user_settings = USettings.objects.all().filter(user=user).first()
            main_shift = main_shifts_served.filter(user=user).first()
            users[user_settings.nickname] = main_shifts_served.filter(user=user).first().id
            name = user_settings.nickname
            index = 1
            for j in range(self.get_object().num_weeks):
                notes.append("")
                for i in range(1, 8):
                    if shift.weeks_data[str(j)][f'R{i}']:
                        served["day" + str(i + (j * 7))] += name
                    if shift.weeks_data[str(j)][f'notes{i}'] != "":
                        notes[j] += name + ": " \
                                     + number_to_day2(index) + " - " + shift.weeks_data[str(j)][f'notes{i}'] + "\n"
            if main_shift.notes != "" and name not in user_notes_added:
                notes_general += name + ": " + main_shift.notes + "\n"
                user_notes_added.append(name)
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ##
        # Calculated Part
        if calculated:
            calc_served = {}
            for x in range(1, self.get_object().num_weeks * 7 + 1):
                day = "day" + str(x)
                calc_served[day] = ""
                split = served[day].split("\n")
                for s in split:
                    if s not in input_days[day + "A"] and s != "" and s != " " and s != "\n":
                        day_before = "day" + str(x - 1)
                        day_after = "day" + str(x + 1)
                        if x != 1 and x != self.get_object().num_weeks * 7:
                            if s not in input_days[day + "M"] and s not in input_days[day + "N"] and \
                                    s not in input_days[day_before + "N"] and s not in input_days[day_after + "M"]:
                                calc_served[day] += s + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_before + "N"] \
                                    and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול בוקר וצהריים) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_before + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק בוקר) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_after + "M"] \
                                    and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק צהריים ולילה) " + "\n"
                            elif s not in input_days[day + "N"] and s not in input_days[day_after + "M"]:
                                calc_served[day] += s + "\n" + " (יכול רק לילה) " + "\n"
                        elif x == 1:
                            if s not in input_days[day + "M"] and s not in input_days[day + "N"] and \
                                    s not in input_days[day_after + "M"]:
                                calc_served[day] += s + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול בוקר וצהריים) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_after + "M"] \
                                    and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק צהריים ולילה) " + "\n"
                            elif s not in input_days[day + "N"] and s not in input_days[day_after + "M"]:
                                calc_served[day] += s + "\n" + " (יכול רק לילה) " + "\n"
                            elif s not in input_days[day + "M"]:
                                calc_served[day] += s + "\n" + " (יכול רק בוקר) " + "\n"
                        else:
                            if s not in input_days[day + "M"] and s not in input_days[day + "N"] and \
                                    s not in input_days[day_before + "N"]:
                                calc_served[day] += s + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_before + "N"] \
                                    and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול בוקר וצהריים) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day_before + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק בוקר) " + "\n"
                            elif s not in input_days[day + "M"] and s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק צהריים ולילה) " + "\n"
                            elif s not in input_days[day + "N"]:
                                calc_served[day] += s + "\n" + " (יכול רק לילה) " + "\n"
            served = calc_served
        ctx["calculated"] = calculated
        ctx["days"] = days
        ctx["served"] = served
        ctx["notes"] = weeks_notes
        ctx["notes_general"] = notes_general
        ctx["num_served"] = len(main_shifts_served)
        ctx["users"] = users
        return ctx

    def get_context_data(self, **kwargs):
        ctx = super(ServedSumReinforcementsDetailView, self).get_context_data(**kwargs)
        calculated = False
        context = self.get_data2(calculated)
        for c in context:
            ctx[c] = context[c]
        return ctx

    def test_func(self):
        return isStaff(self.request.user)

    def post(self, request, *args, **kwargs):
        if request.method == "POST":
            calculated = True
            if 'calculated' in request.POST:
                if request.POST.get("calculated") == "True":
                    calculated = False
                else:
                    calculated = True
            ctx = self.get_data2(calculated)
            return render(request, "mishmar/served_sum_reinforcements.html", ctx)


# View to show staff member shifts served based on organization
class ServedSumShiftDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Organization
    template_name = "mishmar/Served-sum.html"

    def get_context_data(self, **kwargs):
        ctx = super(ServedSumShiftDetailView, self).get_context_data(**kwargs)
        context = get_data(self.get_object())
        for c in context:
            ctx[c] = context[c]
        return ctx

    def test_func(self):
        return isStaff(self.request.user)

    def post(self, request, *args, **kwargs):
        if request.method == "POST":
            if "download" in request.POST:
                ctx = get_data(self.get_object())
                return WriteToExcel(weeks_to_served(ctx["served"]), ctx["notes"], ctx["notes_general"],ctx["days"], self.request.user)


# View to update organizaytion by staff memebers
@user_staff_permission
def organization_update(request, pk=None):
    organization = Organization.objects.all().filter(id=pk).first()
    weeks_shifts = []
    for i in range(organization.num_weeks):
        weeks_shifts.append(organization.weeks_data[str(i)])
    days = []
    shifts = OrganizationShift.objects.all().order_by('shift_num', 'index')
    for x in range(organization.num_weeks * 7):
        days.append(organization.date + datetime.timedelta(days=x))
    if request.method == "POST":
        action = request.POST.get("actions")
        for j in range(organization.num_weeks):
            for i in range(1, 8):
                for key in shifts:
                    organization.weeks_data[str(j)][f"{i}@{key.id}"] = request.POST.get(f"day{i}@{key.id}@{j}", "")
        pub = request.POST.get("pub")
        if pub:
            published = False
        else:
            published = True
        organization.published = published
        organization.save()
        messages.success(request, translate_text(f'עדכון הושלם', request.user, "hebrew"))
        if 'update' == action:
            return HttpResponseRedirect(request.path_info)
        if 'check1' == action:
            organization_valid(organization, request)
            return HttpResponseRedirect(request.path_info)
        elif 'ready' == action:
            return redirect("organization-detail", organization.id)
        elif 'table' == action:
            return redirect("organization-table-shift", organization.id)
        elif 'upload' == action:
            weeks_dicts = uplaod_organize(request, organization)
            temp_weeks = {}
            index = 0
            for week in weeks_dicts:
                temp_weeks[str(index)] = week
                index += 1
            organization.weeks_data = temp_weeks
            organization.save()
            messages.success(request, translate_text(f'העלאה הושלמה', request.user, "hebrew"))
            return HttpResponseRedirect(request.path_info)
        elif 'clear' == action:
            for j in range(organization.num_weeks):
                for i in range(1, 8):
                    for key in shifts:
                        organization.weeks_data[str(j)][f"{i}@{key.id}"] = ""
            organization.save()
            messages.success(request, translate_text(f'איפוס הושלם', request.user, "hebrew"))
            return HttpResponseRedirect(request.path_info)
        elif 'delete' == action:
            Organization.objects.filter(id=organization.id).delete()
            messages.success(request, translate_text(f'מחיקה הושלמה', request.user, "hebrew"))
            return redirect("Served-sum")
    context = {
        "weeks": weeks_shifts,
        "checked": organization.published,
        "organization_id": organization.id,
        "days": days,
        "shifts": shifts,
    }
    return render(request, "mishmar/organization_update.html", context)


# View to show users all the organizations paginated by 1
class OrganizationListView(LoginRequiredMixin, ListView):
    model = Organization
    template_name = "mishmar/organizations_list.html"
    context_object_name = "organizations"
    ordering = ["-date"]
    paginate_by = 1

    def get_context_data(self, **kwargs):
        ctx = super(OrganizationListView, self).get_context_data(**kwargs)
        shifts = OrganizationShift.objects.all().order_by('shift_num', 'index')
        ctx["shifts"] = shifts
        return ctx


def calculate_usage():
    shifts_organization_sum = 0
    shifts_organization_sum += Shift.objects.all().count()
    shifts_organization_sum += Organization.objects.all().count()
    logs_sum = 0
    logs_sum += ValidationLog.objects.all().count()
    logs_sum += Arming_Log.objects.all().count()
    logs_sum += ArmingRequest.objects.all().count()
    events_sum = Event.objects.all().count()
    sum_all = logs_sum + shifts_organization_sum + 1
    sum_all += User.objects.all().count()
    sum_all += Group.objects.all().count()
    sum_all += USettings.objects.all().count()
    sum_all += Gun.objects.all().count()
    sum_all += OrganizationShift.objects.all().count()
    sum_all += Event.objects.all().count()
    sum_all += Post.objects.all().count()
    sum_all += IpBan.objects.all().count()
    return [sum_all, logs_sum, shifts_organization_sum, events_sum]


@user_staff_permission
def data_usage_view(request):
    context = {}
    sum_all, logs_sum, shifts_organization_sum, events_sum = calculate_usage()
    context = {
        "shifts_organization_sum": shifts_organization_sum,
        "logs_sum": logs_sum,
        "sum_all": sum_all,
        "events_sum": events_sum,
    }
    if request.method == "POST":
        if "org" in request.POST:
            return redirect('delete-organization-data')
        elif "log" in request.POST:
            return redirect('delete-logs-data')
        elif 'event' in request.POST:
            return redirect('delete-events-data')
    return render(request, "mishmar/data-usage.html", context)


@user_staff_permission
def delete_organization_data_view(request):
    shifts_organization_sum = 0
    shifts_organization_sum += Shift.objects.all().count()
    shifts_organization_sum += Organization.objects.all().count()
    context = {
        "shifts_organization_sum": shifts_organization_sum,
    }
    if request.method == "POST":
        if 'delete' in request.POST:
            organizations = Organization.objects.all().order_by('-date')
            if len(organizations) > 2:
                exclude_orgs = organizations[:2]
                organizations = organizations[2:]
                for org in organizations:
                    org.delete()
                messages.success(request, "נתונים נמחקו בהצלחה")
                return redirect('data-usage')
            else:
                messages.success(request, "אין מספיק נתונים למחיקה")
                return redirect('data-usage')
        else:
            return redirect('data-usage')
    return render(request, "mishmar/data-organizations-delete.html", context)

@user_staff_permission
def delete_logs_data_view(request):
    logs_sum = 0
    logs_sum += ValidationLog.objects.all().count()
    logs_sum += Arming_Log.objects.all().count()
    logs_sum += ArmingRequest.objects.all().count()
    today = timezone.now()
    month = today.month
    year = today.year
    context = {
        "logs_sum": logs_sum,
        "today": today,
    }
    if request.method == "POST":
        if 'delete' in request.POST:
            Arming_Log.objects.all().exclude(date__year=year, date__month=month).delete()
            ValidationLog.objects.all().exclude(date__year=year, date__month=month).delete()
            messages.success(request, "נתונים נמחקו בהצלחה")
            return redirect('data-usage')
        else:
            return redirect('data-usage')
    return render(request, "mishmar/data-logs-delete.html", context)

@user_staff_permission
def delete_events_data_view(request):
    events_sum = 0
    today = timezone.now()
    events_sum = Event.objects.all().filter(date2__lt=today).count()
    context = {
        "events_sum": events_sum,
        "today": today,
    }
    if request.method == "POST":
        if 'delete' in request.POST:
            Event.objects.all().filter(date2__lt=today).delete()
            messages.success(request, "נתונים נמחקו בהצלחה")
            return redirect('data-usage')
        else:
            return redirect('data-usage')
    return render(request, "mishmar/data-events-delete.html", context)



#TODO
# View to show suggestion organization calculated
class OrganizationSuggestionView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Organization
    template_name = "mishmar/Suggestion.html"

    def get_context_data(self, **kwargs):
        ctx = super(OrganizationSuggestionView, self).get_context_data(**kwargs)
        settings = Settings.objects.all().first()
        try:
            last_organization = Organization.objects.all().order_by('-date')[get_num_organization(self.get_object()) - 1]
        except:
            last_organization = ""
        users = User.objects.all()
        users_settings = USettings.objects.all()
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        ctx["days"] = days
        served, notes = self.get_served()
        if last_organization != "":
            organizer = compare_organizations(served, guards_num, self.get_object(), settings.officer,
                                          last_organization.Day14_2300.split("\n"), users, users_settings)
        else:
            organizer = compare_organizations(served, guards_num, self.get_object(), settings.officer,
                                              "", users, users_settings)
        organized_str = {}
        for key in organizer.organized:
            organized_str[key] = '\n'.join(organizer.organized[key])
        ctx["organized"] = organized_str
        ctx["notes"] = organizer.notes
        ctx["guardsnumbers"] = guards_num
        return ctx

    def post(self, request, *args, **kwargs):
        for x in range(14):
            day = f'M{x}'
            if x == 6 or x == 13:
                guards_num[day] = int(self.request.POST.get(day, 2))
            elif x == 5 or x == 12:
                guards_num[day] = int(self.request.POST.get(day, 0))
            else:
                guards_num[day] = int(self.request.POST.get(day, 5))
            day = f'A{x}'
            if x == 5 or x == 6 or x == 12 or x == 13:
                guards_num[day] = int(self.request.POST.get(day, 0))
            else:
                guards_num[day] = int(self.request.POST.get(day, 3))
            day = f'N{x}'
            if x == 5 or x == 6 or x == 12 or x == 13:
                guards_num[day] = int(self.request.POST.get(day, 2))
            else:
                guards_num[day] = int(self.request.POST.get(day, 1))
        settings = Settings.objects.all().first()
        try:
            last_organization = Organization.objects.all().order_by('-date')[get_num_organization(self.get_object()) - 1]
        except:
            last_organization = ""
        users = User.objects.all()
        users_settings = USettings.objects.all()
        days = []
        for x in range(self.get_object().num_weeks * 7):
            days.append(self.get_object().date + datetime.timedelta(days=x))
        served, notes = self.get_served()
        if last_organization != "":
            organizer = compare_organizations(served, guards_num, self.get_object(), settings.officer,
                                              last_organization.Day14_2300.split("\n"), users, users_settings)
        else:
            organizer = compare_organizations(served, guards_num, self.get_object(), settings.officer,
                                              "", users, users_settings)
        if 'organize' in request.POST:
            return HttpResponseRedirect(self.request.path_info)
        elif 'excel' in request.POST:
            return organizer.WriteToExcel(notes, days, self.request.user)

    def get_served(self):
        served = {}
        for i in range(14):
            served["M" + str(i)] = []
            served["A" + str(i)] = []
            served["N" + str(i)] = []
        shifts_served = Shift.objects.all().filter(date=self.get_object().date)
        notes = {"general": "", "week1": "", "week2": ""}
        for shift in shifts_served:
            user = User.objects.all().filter(username=shift.username).first()
            user_settings = USettings.objects.all().filter(user=user).first()
            name = user_settings.nickname
            shifts = [shift.M1, shift.A1, shift.N1, shift.M2, shift.A2, shift.N2, shift.M3,
                      shift.A3, shift.N3, shift.M4, shift.A4, shift.N4, shift.M5, shift.A5, shift.N5,
                      shift.M6, shift.A6, shift.N6, shift.M7, shift.A7, shift.N7, shift.M8,
                      shift.A8, shift.N8, shift.M9, shift.A9, shift.N9, shift.M10, shift.A10,
                      shift.N10, shift.M11, shift.A11, shift.N11, shift.M12, shift.A12, shift.N12,
                      shift.M13, shift.A13, shift.N13, shift.M14, shift.A14, shift.N14]
            kind = "M"
            count = 0
            index = 0
            for s in shifts:
                if s:
                    served[kind + str(index)].append(user_settings.nickname)
                count = count + 1
                if count == 0:
                    kind = "M"
                elif count == 1:
                    kind = "A"
                elif count == 2:
                    kind = "N"
                else:
                    kind = "M"
                    index = index + 1
                    count = 0
            notes1 = [shift.notes1, shift.notes2, shift.notes3,
                      shift.notes4, shift.notes5, shift.notes6, shift.notes7]
            index = 1
            for n in notes1:
                if n != "":
                    notes["week1"] = notes["week1"] + name + ": " \
                                     + number_to_day2(index) + " - " + n + "\n"
                index += 1
            notes2 = [shift.notes8, shift.notes9, shift.notes10,
                      shift.notes11, shift.notes12, shift.notes13, shift.notes14]
            index = 1
            for n in notes2:
                if n != "":
                    notes["week2"] = notes["week2"] + name + ": " \
                                     + number_to_day2(index) + " - " + n + "\n"
                index += 1
            if shift.notes != "":
                notes["general"] = notes["general"] + name + ": " + shift.notes + "\n"
        return served, notes

    def test_func(self):
        return isStaff(self.request.user)


# Staff member panel View
@user_staff_permission
def staff_panel_view(request):
    settings = Settings.objects.all().first()
    usage = calculate_usage()
    if usage[0] >= 9000:
        messages.error(request, f'כמות נתונים גבוהה מאוד ({usage[0]}) אנא בצע גיבוי ומחיקת נתונים.')
    elif usage[0] >= 8000:
        messages.warning(request, f'כמות נתונים גבוהה ({usage[0]}) אנא בצע גיבוי ומחיקת נתונים.')
    armingrequests = ArmingRequest.objects.all().filter(read=False)
    num_requests = len(armingrequests)
    if num_requests > 0:
        messages.info(request, f'יש {num_requests} בקשות לשינוי ביומן חימוש')
    context = {}
    checked = settings.submitting
    if request.method == 'POST':
        checked = request.POST.get("serv")
        if checked:
            checked = False
        else:
            checked = True
        settings.submitting = checked
        settings.save()
        messages.success(request, "השתנה בהצלחה")
        return redirect('staff-panel')
    context = {
        "checked": checked,
        "today": timezone.now(),
        "year": timezone.now().year,
    }
    return render(request, 'mishmar/staff_panel.html', context)


# Organization shifts design View
@user_staff_permission
def organization_shift_view(request):
    context = {}
    days = []
    for x in range(7):
        days.append(datetime.date.today() + datetime.timedelta(days=x))
    shifts = OrganizationShift.objects.all().order_by('shift_num', 'index')
    if shifts.filter(shift_num=1).count() > 0:
        last_index = shifts.filter(shift_num=1).order_by('index').last().index + 1
    else:
        last_index = 1
    context = {
        "shifts": shifts,
        "days": days,
        "last_index": last_index
    }
    if request.method == 'POST':
        if 'add' in request.POST or 'change' in request.POST:
            if 'add' in request.POST:
                shift = OrganizationShift()
                shift_id = ""
            else:
                shift_id = request.POST.get('change')
                shift = OrganizationShift.objects.all().filter(id=shift_id).first()
            shift.shift_num = request.POST.get(f"shift_num{shift_id}")
            shift.index = int(request.POST.get(f"index_num{shift_id}"))
            shift.title = request.POST.get(f"title{shift_id}")
            shift.sub_title = request.POST.get(f"sub_title{shift_id}")
            shift.opening = checkbox(request.POST.get(f"opening{shift_id}"))
            shift.manager = checkbox(request.POST.get(f"manager{shift_id}"))
            shift.pull = checkbox(request.POST.get(f"pull{shift_id}"))
            if len(OrganizationShift.objects.filter(index=shift.index, shift_num=shift.shift_num)) > 0:
                shifts = OrganizationShift.objects.filter(index__gte=shift.index, shift_num=shift.shift_num)
                index_temp = int(shift.index)
                for s in shifts:
                    if s.id != shift.id:
                        s.index = index_temp + 1
                        s.save()
                        index_temp += 1
            shift.save()
            if 'add' in request.POST:
                messages.success(request, "נוסף בהצלחה")
            else:
                messages.success(request, "נשמר בהצלחה")
            return HttpResponseRedirect(request.path_info)
        elif 'delete' in request.POST:
            shift_id = request.POST.get('delete')
            shift = OrganizationShift.objects.all().filter(id=shift_id).first()
            if len(OrganizationShift.objects.filter(index=shift.index, shift_num=shift.shift_num)) > 0:
                shifts = OrganizationShift.objects.filter(index__gte=shift.index, shift_num=shift.shift_num)
                for s in shifts:
                    if s.id != shift.id:
                        s.index = int(s.index) - 1
                        s.save()
            shift.delete()
            messages.success(request, "נמחק בהצלחה")
            return HttpResponseRedirect(request.path_info)
        elif 'view_org' in request.POST:
            return render(request, 'mishmar/organization-ready.html', context)
        elif 'back' in request.POST:
            return render(request, 'mishmar/organization_shift.html', context)
    return render(request, 'mishmar/organization_shift.html', context)
    
class PostCreateView(CreateView, UserPassesTestMixin, LoginRequiredMixin):
    model = Post
    template_name = "mishmar/post-new.html"
    fields = ("username", "date", "title", "text")

    def get_context_data(self, **kwargs):
        ctx = super(PostCreateView, self).get_context_data(**kwargs)
        ctx["today"] = timezone.now()
        ctx["view_type"] = "create"
        return ctx

    def test_func(self):
        return isStaff(self.request.user)
    
    def post(self, request, *args, **kwargs):
        new_post = Post()
        new_post.date = timezone.now()
        new_post.username = self.request.user
        new_post.text = self.request.POST.get('text')
        new_post.title = self.request.POST.get('title')
        new_post.save()
        messages.success(request, "פוסט נוצר")
        return redirect("post-list")

class PostUpdateView(UpdateView, UserPassesTestMixin, LoginRequiredMixin):
    model = Post
    template_name = "mishmar/post-new.html"
    fields = ("title", "text")
    context_object_name = "form"

    def get_context_data(self, **kwargs):
        ctx = super(PostUpdateView, self).get_context_data(**kwargs)
        ctx["today"] = timezone.now()
        ctx["view_type"] = "update"
        ctx["text"] = self.get_object().text
        return ctx

    def test_func(self):
        return isStaff(self.request.user)
    
    def post(self, request, *args, **kwargs):
        if "delete" in request.POST:
            self.get_object().delete()
            messages.success(request, "נמחק בהצלחה")
            return redirect("post-list")
        else:
            post = Post.objects.all().filter(id=self.get_object().id).first()
            post.date = timezone.now()
            post.title = request.POST.get("title")
            post.text = request.POST.get("text")
            post.save()
            messages.success(request, 'עודכן בהצלחה')
            return redirect("post-list")

class PostListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Post
    template_name = "mishmar/post-list.html"
    context_object_name = "posts"
    paginate_by = 3
    ordering = ["-date"]

    def test_func(self):
        return isStaff(self.request.user)

class EventCreateView(CreateView, UserPassesTestMixin, LoginRequiredMixin):
    model = Event
    template_name = "mishmar/event-new.html"
    fields = ("nickname", "date2", "description")

    def get_context_data(self, **kwargs):
        ctx = super(EventCreateView, self).get_context_data(**kwargs)
        users = User.objects.all().exclude(username="metagber").exclude(username="admin")
        names = []
        for user in users:
            u_settings = USettings.objects.all().filter(user=user).first()
            names.append(u_settings.nickname)
        ctx["today"] = timezone.now()
        ctx["view_type"] = "create"
        ctx["names"] = names
        return ctx

    def test_func(self):
        return isStaff(self.request.user)
    
    def post(self, request, *args, **kwargs):
        new_event = Event()
        new_event.date2 = self.request.POST.get('date')
        new_event.nickname = self.request.POST.get('names')
        new_event.description = self.request.POST.get('description')
        new_event.save()
        messages.success(request, "אירוע נוצר")
        return redirect("event-list")

class EventUpdateView(UpdateView, UserPassesTestMixin, LoginRequiredMixin):
    model = Event
    template_name = "mishmar/event-new.html"
    fields = ("nickname", "date2", "description")
    context_object_name = "event"

    def get_context_data(self, **kwargs):
        ctx = super(EventCreateView, self).get_context_data(**kwargs)
        users = User.objects.all().exclude(username="metagber").exclude(username="admin")
        names = []
        for user in users:
            u_settings = USettings.objects.all().filter(user=user).first()
            names.append(u_settings.nickname)
        ctx["today"] = timezone.now()
        ctx["view_type"] = "update"
        ctx["names"] = names
        return ctx

    def test_func(self):
        return isStaff(self.request.user)
    
    def post(self, request, *args, **kwargs):
        if 'delete' in request.POST:
            self.get_object().delete()
            messages.success(request, "אירוע נמחק")
            return redirect("event-list")
        else:
            event = Event.objects.all().filter(id=self.get_object().id).first()
            event.date2 = self.request.POST.get('date')
            event.nickname = self.request.POST.get('names')
            event.description = self.request.POST.get('description')
            event.save()
            messages.success(request, "אירוע נוצר")
            return redirect("event-list")

class EventListView(UserPassesTestMixin, LoginRequiredMixin, ListView):
    model = Event
    template_name = "mishmar/event-list.html"
    context_object_name = "events"
    paginate_by = 10
    ordering = ["-date2"]

    def get_context_data(self, **kwargs):
        ctx = super(EventListView, self).get_context_data(**kwargs)
        users = User.objects.all().exclude(username="metagber").exclude(username="admin")
        names = []
        for user in users:
            u_settings = USettings.objects.all().filter(user=user).first()
            names.append(u_settings.nickname)
        ctx["today"] = timezone.now()
        ctx["names"] = names
        return ctx

    def test_func(self):
        return isStaff(self.request.user)
    
    def post(self, request, *args, **kwargs):
        if "new" in request.POST:
            new_event = Event()
            new_event.date2 = self.request.POST.get('date')
            new_event.nickname = self.request.POST.get('names')
            new_event.description = self.request.POST.get('description')
            new_event.save()
            messages.success(request, "אירוע נוצר")
        elif 'delete' in request.POST:
            id = request.POST.get("delete")
            event = Event.objects.all().filter(id=id).first()
            event.delete()
            messages.success(request, "אירוע נמחק")
        else:
            id = request.POST.get("change")
            event = Event.objects.all().filter(id=id).first()
            event.date2 = self.request.POST.get(f'date{id}')
            event.nickname = self.request.POST.get(f'names{id}')
            event.description = self.request.POST.get(f'description{id}')
            event.save()
            messages.success(request, "אירוע שונה")
        return HttpResponseRedirect(request.path_info)

class IpBanListView(UserPassesTestMixin, LoginRequiredMixin, ListView):
    model = IpBan
    template_name = "mishmar/ipban-list.html"
    context_object_name = "ips"
    paginate_by = 10
    ordering = ["num_tries"]

    def get_context_data(self, **kwargs):
        ctx = super(IpBanListView, self).get_context_data(**kwargs)
        return ctx

    def test_func(self):
        return isStaff(self.request.user)
    
    def post(self, request, *args, **kwargs):
        if 'delete' in request.POST:
            id = request.POST.get('delete')
            ip = IpBan.objects.all().filter(id=id).first()
            ip.delete()
            messages.success(request, "נמחק בהצלחה")
        else:
            id = request.POST.get('change')
            ip = IpBan.objects.all().filter(id=id).first()
            ip.num_tries = request.POST.get(f'tries{id}')
            ip.save()
            messages.success(request, "נשמר בהצלחה")
        return HttpResponseRedirect(request.path_info)

# Side functions

# get shift served data
def get_data(object):
    ctx = {}
    served = {}
    counters = {}
    settings = Settings.objects.all().first()
    minimum_day_morning = 6
    minimum_day_noon = 6
    if settings.friday_morning:
        minimum_day_morning = 7
    if settings.friday_noon:
        minimum_day_noon = 7
    for j in range(object.num_weeks):
        served[str(j)] = {}
        for i in range(1, 8):
            served[str(j)]["M" + str(i)] = ""
            served[str(j)]["A" + str(i)] = ""
            served[str(j)]["N" + str(i)] = ""
    main_shifts_served = Shift.objects.all().filter(organization=object)
    weeks_notes = []
    for i in range(object.num_weeks):
        weeks_notes.append("")
    notes_general = ""
    user_notes_added = []
    users = {}
    shifts_keys = []
    for i in range(1, 8):
        shifts_keys.append("M" + str(i))
        shifts_keys.append("A" + str(i))
        shifts_keys.append("N" + str(i))
    for shift in main_shifts_served:
        user = shift.user
        user_settings = USettings.objects.all().filter(user=user).first()
        users[user_settings.nickname] = shift.id
        name = user_settings.nickname
        name = name.replace("\n", "")
        name = name.replace("\r", "")
        morning = False
        for i in range(object.num_weeks):
            counters[f"M-{name}-{i}"] = 0
            counters[f"A-{name}-{i}"] = 0
            for j in range(1, 8):
                if shift.weeks_data[str(i)][f'M{j}']:
                    served[str(i)][f'M{j}'] += name
                    morning = True
                    if j < minimum_day_morning:
                        counters[f"M-{name}-{i}"] += 1
                if not shift.weeks_data[str(i)][f'P{j}']:
                    if morning:
                        served[str(i)][f'M{j}'] += "\n" + "(לא משיכה)" + "\n"
                    else:
                        served[str(i)][f'M{j}'] += "\n"
                else:
                    if morning:
                        served[str(i)][f'M{j}'] += "\n"
                morning = False
                if shift.weeks_data[str(i)][f'A{j}']:
                    served[str(i)][f'A{j}'] += name + "\n"
                    if j < minimum_day_noon:
                        counters[f"A-{name}-{i}"] += 1
                if shift.weeks_data[str(i)][f'N{j}']:
                    served[str(i)][f'N{j}'] += name + "\n"
                if shift.weeks_data[str(i)][f'notes{j}'] != "":
                    weeks_notes[i] += name + ": " + number_to_day2(j) + " - " + shift.weeks_data[str(i)][f'notes{j}'] + "\n"
        if shift.notes != "" and name not in user_notes_added:
            notes_general += name + ": " + shift.notes + "\n"
            user_notes_added.append(name)
    not_qual_users = {}
    for key in counters:
        shift, name, week = key.split("-")
        if shift == 'M':
            if counters[key] < 2:
                not_qual_users[name] = users[name]
        else:
            if counters[key] < 1:
                not_qual_users[name] = users[name]
    days = []
    for x in range(object.num_weeks * 7):
        days.append(object.date + datetime.timedelta(days=x))
    ctx["days"] = days
    ctx["served"] = served
    ctx["notes"] = weeks_notes
    ctx["notes_general"] = notes_general
    ctx["num_served"] = len(main_shifts_served)
    ctx["users"] = users
    ctx["not_qual_users"] = not_qual_users
    ctx["not_qual_num"] = len(not_qual_users.keys())
    return ctx


# Day number to string name
def number_to_day2(num):
    day = "יום "
    days = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
    return day + days[num - 1]

# Get context for select inputs for arming views
def put_arming_context(ctx):
    settings = Settings.objects.all().first()
    num_mags_list = range(1, settings.num_mags + 1)
    hand_cuffs_list = range(1, settings.hand_cuffs + 1)
    mag_case_list = range(1, settings.num_mag_cases + 1)
    gun_case_list = range(1, settings.num_gun_cases + 1)
    ctx["num_mags_list"] = num_mags_list
    ctx["hand_cuffs_list"] = hand_cuffs_list
    ctx["mag_case_list"] = mag_case_list
    ctx["gun_case_list"] = gun_case_list
    ctx["guns"] = Gun.objects.all()
    return ctx

# return boolean values for checkboxes
def checkbox(value):
    if value == "on":
        return True
    return False

# Check if the user has already submitted shifts
def already_submitted(user):
    organization = Organization.objects.order_by('-date').first()
    shifts = Shift.objects.filter(organization=organization)
    if len(shifts) == 0:
        return False
    else:
        if len(shifts.filter(user=user)) == 0:
            return False
    return True


# Get organization data and return as dictionary
def get_input(organization):
    weeks = []
    for i in range(organization.num_weeks):
        weeks.append(organization.weeks_data[str(i)])
    index = 0
    new_dict = {}
    for week in weeks:
        if index != 0:
            for key in week.keys():
                split = key.split('@')
                day = int(split[0]) + index * 7
                new_dict[f'{day}@{split[1]}'] = week[key]
        else:
            new_dict = week
        index += 1
    return new_dict

# Extract data from excel and return it as a list of variables
def extract_data(request, organization):
    myfile = request.FILES['myfile']
    file_object = myfile.file
    wb = openpyxl.load_workbook(file_object)
    sheet = wb.active
    ## Green background FFC6EFCE
    ## Green font FF006100
    ## Red background FFFFC7CE
    ## Red font with red background FF9C0006
    ## White background 00000000
    ## Black font Values must be of type <class 'str'> rgb=None, indexed=None, auto=None, theme=1, tint=0.0, type='theme'
    ## Red font FFFF0000
    ## Empty font Values must be of type <class 'str'> rgb=None, indexed=None, auto=None, theme=1, tint=0.0, type='theme'
    ## Empty value None
    ## orange background FFFFEB9C
    # str(sheet.cell(10, 3).value)
    # str(sheet.cell(10, 3).font.color.rgb)  # Get the font color in the table
    # str(sheet.cell(10, 3).fill.fgColor.rgb) # background color
    # if str(sheet.cell(10, 3).font.color.rgb) == "Values must be of type <class 'str'>":
    end_morning_str = ""
    end_noon_str = ""
    end_night_str = ""
    for rng in sheet.merged_cells.ranges:
        if 'A5' in rng:
            end_morning_str = str(rng)
            break
    end_morning_str = end_morning_str.replace("A5:A", "")
    end_morning = int(end_morning_str)
    for rng in sheet.merged_cells.ranges:
        if f'A{end_morning + 1}' in rng:
            end_noon_str = str(rng)
            break
    end_noon_str = end_noon_str.replace(f'A{end_morning + 1}:A', "")
    end_noon = int(end_noon_str)
    for rng in sheet.merged_cells.ranges:
        if f'A{end_noon + 1}' in rng:
            end_night_str = str(rng)
            break
    end_night_str = end_night_str.replace(f'A{end_noon + 1}:A', "")
    end_night = int(end_night_str)
    col = 1
    names_days = {}
    no_pull_names = {}
    for x in range(organization.num_weeks * 7):
        col += 1
        names_days[f'day{x}_morning'] = []
        no_pull_names[f'day{x}'] = []
        names_days[f'day{x}_noon'] = []
        names_days[f'day{x}_night'] = []
        for j in range(5, end_morning + 1):
            if str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFC6EFCE' \
                    or str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFFFEB9C':
                names_days[f'day{x}_morning'].append(str(sheet.cell(j, col).value))
                if str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFFFEB9C':
                    no_pull_names[f'day{x}'].append(str(sheet.cell(j, col).value))
        for j in range(end_morning + 1, end_noon + 1):
            if str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFC6EFCE':
                names_days[f'day{x}_noon'].append(str(sheet.cell(j, col).value))
        for j in range(end_noon + 1, end_night + 1):
            if str(sheet.cell(j, col).fill.fgColor.rgb) == 'FFC6EFCE':
                names_days[f'day{x}_night'].append(str(sheet.cell(j, col).value))
    # extract from database
    shifts = Shift.objects.all().filter(organization=organization)
    settings = Settings.objects.all().first()
    max_seq0 = settings.max_seq0
    max_seq1 = settings.max_seq1
    sequence_count = {}
    max_out_names = [[], []]
    for s in shifts:
        user = s.user
        user_settings = USettings.objects.all().filter(user=user).first()
        name = user_settings.nickname
        sequence_count[f'{name}0'] = s.seq_night
        sequence_count[f'{name}1'] = s.seq_noon
        if sequence_count[f'{name}0'] >= max_seq0:
            max_out_names[0].append(name)
        if sequence_count[f'{name}1'] >= max_seq1:
            max_out_names[1].append(name)
    return [names_days, no_pull_names, sequence_count, max_out_names, max_seq0, max_seq1]


# Given data from excel and puts it in the organization database based on hours
def uplaod_organize(request, organization):
    extracted_data = extract_data(request, organization)
    names_days = extracted_data[0]
    no_pull_names = extracted_data[1]
    sequence_count = extracted_data[2]
    max_out_names = extracted_data[3]
    max_seq0 = extracted_data[4]
    max_seq1 = extracted_data[5]
    shift_keys = OrganizationShift.objects.all()
    morning_keys = shift_keys.filter(shift_num=1)
    noon_keys = shift_keys.filter(shift_num=2)
    night_keys = shift_keys.filter(shift_num=3)
    try:
        before_organization = Organization.objects.order_by('-date')[1]
    except:
        before_organization = None
    if before_organization is not None:
        morning_names = []
        for key in morning_keys:
            if f'7@{key.id}' in before_organization.weeks_data[str(before_organization.num_weeks - 1)].keys():
                morning_names +=  before_organization.weeks_data[str(before_organization.num_weeks - 1)][f'7@{key.id}'].split("\n")
        noon_names = []
        for key in noon_keys:
            if f'7@{key.id}' in before_organization.weeks_data[str(before_organization.num_weeks - 1)].keys():
                noon_names += before_organization.weeks_data[str(before_organization.num_weeks - 1)][f'7@{key.id}'].split("\n")
        night_names = []
        for key in night_keys:
            if f'7@{key.id}' in before_organization.weeks_data[str(before_organization.num_weeks - 1)].keys():
                night_names +=  before_organization.weeks_data[str(before_organization.num_weeks - 1)][f'7@{key.id}'].split("\n")
        before_names = {"motsash": night_names, "noon": noon_names, "morning": morning_names}
        for key in before_names:
            for v in before_names[key]:
                if v == '' or v == '\r' or v == ' ':
                    count = before_names[key].count(v)
                    for x in range(count):
                        before_names[key].remove(v)
                else:
                    before_names[key][before_names[key].index(v)] = v.replace("\r", "")
    else:
        before_names = {"motsash": [], "noon": [], "morning": []}
    manager_group = Group.objects.filter(name="manager").first()
    manager_group_users = User.objects.filter(groups=manager_group)
    manager_fields = morning_keys.filter(manager=True)
    pull_fields = morning_keys.filter(pull=True)
    opening_fields = morning_keys.filter(opening=True)
    morning_keys = morning_keys.filter(manager=False, pull=False, opening=False)
    managers = []
    for m in manager_group_users:
        user_settings = USettings.objects.all().filter(user=m).first()
        managers.append(user_settings.nickname)
    num_week = organization.num_weeks - 1
    days_count = 0
    weeks_dicts = []
    for i in range(organization.num_weeks):
        weeks_dicts.append({})
    for i in range(organization.num_weeks):
        fields = organization.weeks_data[str(0)].keys()
        for field in fields:
            weeks_dicts[i][field] = ""
    for i in range(organization.num_weeks * 7 - 1, -1, -1):
        names_x = i
        x = i - 7 * num_week + 1
        manager_shift_list_put = []
        count_manager = 0
        opening_shift_list_put = []
        noon_shift_list_put = []
        if x != 6 and x != 7:
            is_manager = False
            for name in managers:
                if name in names_days[f'day{names_x}_morning']:
                    for key in manager_fields:
                        if key not in manager_shift_list_put:
                            weeks_dicts[num_week][f'{x}@{key.id}'] = name
                            names_days[f'day{names_x}_morning'].remove(name)
                            break
                    count_manager += 1
                    if count_manager == len(manager_fields):
                        is_manager = True
                        break
            if not is_manager:
                if Settings.objects.last().officer in names_days[f'day{names_x}_morning']:
                    for key in manager_fields:
                        if key not in manager_shift_list_put:
                            weeks_dicts[num_week][f'{x}@{key.id}'] = Settings.objects.last().officer
                            names_days[f'day{names_x}_morning'].remove(Settings.objects.last().officer)
            temp_morning = []
            if len(names_days[f'day{names_x}_morning']) > 0:
                for name in names_days[f'day{names_x}_morning']:
                    temp_morning.append(name)
                for name in temp_morning:
                    if name in no_pull_names[f'day{names_x}'] or\
                            name in names_days[f'day{names_x}_night']:
                        if x == 1:
                            temp_morning.remove(name)
                        elif name in names_days[f'day{names_x - 1}_noon']:
                            temp_morning.remove(name)
                if len(temp_morning) > 0:
                    for key in pull_fields:
                        inserted = insert_random(weeks_dicts[num_week], temp_morning, f'{key.id}', x, 0)
                        names_days[f'day{names_x}_morning'].remove(inserted)
                else:
                    temp_morning = []
                    for name in names_days[f'day{names_x}_morning']:
                        temp_morning.append(name)
                    for name in temp_morning:
                        if x != 1:
                            if name in names_days[f'day{names_x - 1}_noon']:
                                temp_morning.remove(name)
                        else:
                            if name in before_names["noon"] or name in before_names["morning"]:
                                temp_morning.remove(name)
                    if len(temp_morning) > 0:
                        for key in pull_fields:
                            inserted = insert_random(weeks_dicts[num_week], temp_morning, f'{key.id}', x, 0)
                            names_days[f'day{names_x}_morning'].remove(inserted)
                            temp_morning.remove(inserted)
            chosen = False
            if x == 1 and num_week == 0:
                for key in opening_fields:
                    chosen = search_and_put(weeks_dicts[num_week], before_names["noon"], names_days[f'day{names_x}_morning'], x,
                                                f'{key.id}', max_out_names[0], 0, sequence_count, max_seq0,
                                                max_seq1, True, [])
                    if chosen:
                        opening_shift_list_put.append(f'{key.id}')
                if not chosen:
                    for key in opening_fields:
                        if key not in opening_shift_list_put:
                            chosen = search_and_put(weeks_dicts[num_week], before_names["morning"], names_days[f'day{names_x}_morning'], x,
                                                        f'{key.id}', max_out_names[0], 0, sequence_count, max_seq0,
                                                        max_seq1, True, [])
                            if chosen:
                                opening_shift_list_put.append(f'{key.id}')
                if not chosen:
                    temp_morning = seperate_list(names_days[f'day{names_x}_morning'], max_out_names)
                    if len(temp_morning) > 0:
                        for key in opening_fields:
                            if key not in opening_shift_list_put:
                                chosen = insert_random(weeks_dicts[num_week], temp_morning, f'{key.id}', x, 0)
                                if chosen is not None:
                                    names_days[f'day{names_x}_morning'].remove(chosen)
                                    opening_shift_list_put.append(f'{key.id}')
                    else:
                        for key in opening_fields:
                            if key not in opening_shift_list_put:
                                insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], f'{key.id}', x, 0)
                # noon
                index = 0
                for key in noon_keys:
                    if index < len(noon_keys) - 1:
                        insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], f'{key.id}', x, 0)
                    else:
                        insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], x, f'{key.id}')
            # morning opening
            else:
                if x > 2:
                    temp_morning = []
                    for name in names_days[f'day{names_x}_morning']:
                        temp_morning.append(name)
                    for key in opening_fields:
                        chosen = search_and_put(weeks_dicts[num_week], names_days[f'day{names_x - 1}_noon'], names_days[f'day{names_x}_morning']
                                                    , x, f'{key.id}', max_out_names[1], 1, sequence_count, max_seq0,
                                                    max_seq1, True, names_days[f'day{names_x - 2}_night'])
                        if chosen:
                            opening_shift_list_put.append(f'{key.id}')
                else:
                    for key in opening_fields:
                        chosen = search_and_put(weeks_dicts[num_week], names_days[f'day{names_x - 1}_noon'],
                                                    names_days[f'day{names_x}_morning'], x,
                                                    f'{key.id}', max_out_names[1], 1, sequence_count, max_seq0,
                                                    max_seq1, True, [])
                        if chosen:
                            opening_shift_list_put.append(f'{key.id}')
                if not chosen:
                    temp_morning = seperate_list(names_days[f'day{names_x}_morning'], max_out_names)
                    if len(temp_morning) > 0:
                        for key in opening_fields:
                            if key not in opening_shift_list_put:
                                chosen = insert_random(weeks_dicts[num_week], temp_morning, f'{key.id}', x, 0)
                                if chosen is not None:
                                    names_days[f'day{names_x}_morning'].remove(chosen)
                                    opening_shift_list_put.append(f'{key.id}')
                    else:
                        for key in opening_fields:
                            if key not in opening_shift_list_put:
                                insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], f'{key.id}', x, 0)       
                # noon
                index = 0
                for key in noon_keys:
                    if index < len(noon_keys) - 1:
                        insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], f'{key.id}', x, 0)
                    else:
                        insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], x, f'{key.id}')
                    index += 1
            # morning
            index = 0
            for key in morning_keys:
                if index < len(morning_keys) - 1:
                    insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], f'{key.id}', x, 0)
                else:
                    insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], x, f'{key.id}')
                index += 1
            # night
            index = 0
            for key in night_keys:
                if index < len(night_keys) - 1:
                    insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_night'], f'{key.id}', x, 0)
                else:
                    insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_night'], x, f'{key.id}')
                index += 1
        else:
            # noon
            index = 0
            for key in noon_keys:
                if index < len(noon_keys) - 1:
                    insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], f'{key.id}', x, 0)
                else:
                    insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_noon'], x, f'{key.id}')
                index += 1
            # morning
            index = 0
            for key in morning_keys:
                if index < len(morning_keys) - 1:
                    insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], f'{key.id}', x, 0)
                else:
                    insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_morning'], x, f'{key.id}')
                index += 1
            # night
            index = 0
            for key in night_keys:
                if index < len(night_keys) - 1:
                    insert_random(weeks_dicts[num_week], names_days[f'day{names_x}_night'], f'{key.id}', x, 0)
                else:
                    insert_all_to_form(weeks_dicts[num_week], names_days[f'day{names_x}_night'], x, f'{key.id}')
                index += 1
        if days_count == 6:
            num_week -= 1
            days_count = -1
        days_count += 1
    return weeks_dicts

# Searches free space to put user 
def search_and_put(weeks_dict, for_list, check_list, day, time, max_out_names, seq,
                   sequence_count, max_seq0, max_seq1, is_seq, extra_seq):
    if seq == 0:
        max_seq = max_seq0
    else:
        max_seq = max_seq1
    for name in for_list:
        if name in check_list and name not in extra_seq:
            if is_seq:
                if name not in max_out_names:
                    if f'{name}{seq}' in sequence_count.keys():
                        sequence_count[f'{name}{seq}'] += 1
                        if sequence_count[f'{name}{seq}'] >= max_seq:
                            max_out_names.append(name)
                    if weeks_dict.get(f'{day}@{time}', "") == "":
                        weeks_dict[f'{day}@{time}'] = name
                    else:
                        weeks_dict[f'{day}@{time}'] += "\n" + name
                    check_list.remove(name)
                    return True
            else:
                if weeks_dict.get(f'{day}@{time}', "") == "":
                    weeks_dict[f'{day}@{time}'] = name
                else:
                    weeks_dict[f'{day}@{time}'] += "\n" + name
                check_list.remove(name)
                return True
    return False


# insert data to form
def insert_all_to_form(weeks_dict, for_list, day, time):
    weeks_dict[f'{day}@{time}'] = '\n'.join(for_list)
    for name in for_list:
        for_list.remove(name)

# seperate user that reached max sequence count
def seperate_list(shift, max_out_names):
    new_list = []
    for s in shift:
        if s not in max_out_names:
            new_list.append(s)
    return new_list

# insert random user
def insert_random(weeks_dict, list1, time, day, count):
    if len(list1) > 0:
        r = random.randint(0, len(list1) - 1)
        if count == 0:
            weeks_dict[f'{day}@{time}'] = list1[r]
        else:
            weeks_dict[f'{day}@{time}'] += "\n" + list1[r]
        return list1.pop(r)
    else:
        return None

# convert day to string day and week
def num_to_day_and_week(num):
    days = ["שבת", "ראשון", "שני", "שלישי","רביעי", "חמישי", "שישי"]
    day = num % 7
    day = days[day]
    if num % 7 == 0:
        week = int(num / 7)
    else:
        week = int(num / 7) + 1
    return f'{day} בשבוע ה-{week}'

# Checks if organization is valid
def organization_valid(organization, request):
    organization1 = get_input(organization)
    input_days = {}
    valid = True
    morning_keys = OrganizationShift.objects.all().filter(shift_num=1)
    noon_keys = OrganizationShift.objects.all().filter(shift_num=2)
    night_keys = OrganizationShift.objects.all().filter(shift_num=3)
    for i in range(1, organization.num_weeks * 7 + 1):
        day = "day" + str(i)
        input_days[day + "M"] = []
        input_days[day + "A"] = []
        input_days[day + "N"] = []
        for key in morning_keys:
            input_days[day + "M"] += organization1[f'{i}@{key.id}'].split("\n")
        for key in noon_keys:
            input_days[day + "A"] += organization1[f'{i}@{key.id}'].split("\n")
        for key in night_keys:
            input_days[day + "N"] += organization1[f'{i}@{key.id}'].split("\n")
    for key in input_days:
        for i in range(len(input_days[key])):
            input_days[key][i] = input_days[key][i].replace(" ", "")
            input_days[key][i] = input_days[key][i].replace("\n", "")
            input_days[key][i] = input_days[key][i].replace("\r", "")
    for key in input_days:
        for i in range(input_days[key].count('')):
            input_days[key].remove('')
    massages_sent = []
    for key in input_days:
        for name in input_days[key]:
            num_day = key.replace("day", "")
            num_day = num_day.replace("A", "")
            num_day = num_day.replace("N", "")
            num_day = num_day.replace("M", "")
            message1 = name + " ביום " + num_to_day_and_week(int(num_day)) + " בשתי משמרות רצופות"
            message2 = name + " ביום " + num_to_day_and_week(int(num_day)) + " באותה משמרת פעמיים"
            day = "day" + num_day
            day_before = "day" + str(int(num_day) - 1)
            day_after = "day" + str(int(num_day) + 1)
            if name in input_days[day + "M"]:
                if is_more_than_once(input_days[day + "M"], name):
                    if message2 not in massages_sent:
                        messages.info(request, translate_text(message2, request.user, "hebrew"))
                        massages_sent.append(message2)
                    valid = False
                if int(num_day) != 1:
                    if name in input_days[day + "A"] or name in input_days[day_before + "N"]:
                        if message1 not in massages_sent:
                            messages.info(request, translate_text(message1, request.user, "hebrew"))
                            massages_sent.append(message1)
                        valid = False
                else:
                    if name in input_days[day + "A"]:
                        if message1 not in massages_sent:
                            messages.info(request, translate_text(message1, request.user, "hebrew"))
                            massages_sent.append(message1)
                        valid = False
            if name in input_days[day + "A"]:
                if is_more_than_once(input_days[day + "A"], name):
                    if message2 not in massages_sent:
                        messages.info(request, translate_text(message2, request.user, "hebrew"))
                        massages_sent.append(message2)
                    valid = False
                if name in input_days[day + "M"] or name in input_days[day + "N"]:
                    if message1 not in massages_sent:
                        messages.info(request, translate_text(message1, request.user, "hebrew"))
                        massages_sent.append(message1)
                    valid = False
            if name in input_days[day + "N"]:
                if is_more_than_once(input_days[day + "N"], name):
                    if message2 not in massages_sent:
                        messages.info(request, translate_text(message2, request.user, "hebrew"))
                        massages_sent.append(message2)
                    valid = False
                if int(num_day) != organization.num_weeks * 7:
                    if name in input_days[day + "A"] or name in input_days[day_after + "M"]:
                        if message1 not in massages_sent:
                            messages.info(request, translate_text(message1, request.user, "hebrew"))
                            massages_sent.append(message1)
                        valid = False
                else:
                    if name in input_days[day + "A"]:
                        if message1 not in massages_sent:
                            messages.info(request, translate_text(message1, request.user, "hebrew"))
                            massages_sent.append(message1)
                        valid = False
    if valid:
        messages.success(request, translate_text("סידור תקין", request.user, "hebrew"))

# Check if name apears more than once in list
def is_more_than_once(list, name):
    num = 0
    for n in list:
        if n == name:
            num += 1
    if num > 1:
        return True
    return False

# turn array of weeks to full dict
def weeks_to_served(weeks):
    served = {}
    for j in range(len(weeks)):
        for i in range(1, 8):
            new_day = i + (7 * j)
            served[f'M{new_day}'] = weeks[str(j)][f'M{i}']
            served[f'A{new_day}'] = weeks[str(j)][f'A{i}']
            served[f'N{new_day}'] = weeks[str(j)][f'N{i}']
    return served

# Write shifts served data to excel
def WriteToExcel(served, notes, notes_general, dates, user):
    # Create a workbook and add a worksheet.
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()
    worksheet.right_to_left()

    days = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]
    user_settings = USettings.objects.all().filter(user=user).first()
    if user_settings.language == 'english':
        days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    maxes = {"morning": 0, "after_noon": 0, "night": 0}

    for key in served:
        temp = 0
        if key.count("M"):
            split = served[key].split("\n")
            for x in range(len(split)):
                if split[x] != "(לא משיכה)":
                    temp += 1
            if temp > maxes["morning"]:
                maxes["morning"] = temp
        elif key.count("A"):
            split = served[key].split("\n")
            if len(split) > maxes["after_noon"]:
                maxes["after_noon"] = len(split)
        else:
            split = served[key].split("\n")
            if len(split) > maxes["night"]:
                maxes["night"] = len(split)

    maxes["morning"] += 1
    maxes["after_noon"] += 2
    maxes["night"] += 2

    # Write a total using a formula.
    title_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 24,
        'fg_color': 'white'})
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
    })
    cell_no_pull_format = workbook.add_format({
        'font_color': "#ff0000"
    })
    border_bottom_format = workbook.add_format({
        'bottom': 5,
        'bottom_color': '#000000'
    })
    border_left_format = workbook.add_format({
        'left': 5,
        'left_color': '#000000'
    })
    border_left_bottom_format = workbook.add_format({
        'left': 5,
        'left_color': '#000000',
        'bottom': 5,
        'bottom_color': '#000000',
    })
    border_right_format = workbook.add_format({
        'right': 5,
        'right_color': '#000000'
    })
    border_right_bottom_format = workbook.add_format({
        'right': 5,
        'right_color': '#000000',
        'bottom': 5,
        'bottom_color': '#000000',
    })
    # Building first Structure
    col = 0
    for x in range(len(dates) + 1):
        worksheet.write(4 + maxes["morning"], col, None, border_bottom_format)
        worksheet.write(4 + maxes["after_noon"] + maxes["morning"], col, None, border_bottom_format)
        col += 1
    row = 0
    sum_maxes = maxes["morning"] + maxes["after_noon"] + maxes["night"] + 6
    for i in range(int(len(dates) / 7)):
        row = 0
        for x in range(sum_maxes):
            if x == 4 + maxes["morning"] or x == 4 + maxes["morning"] + maxes["after_noon"]:
                worksheet.write(row, 7 + 7 * i, None, border_right_bottom_format)
            else:
                worksheet.write(row, 7 + 7 * i, None, border_right_format)
            row += 1
    # Building second Structure
    worksheet.merge_range('A1:H2', translate_text('הגשות', user, "hebrew"), title_format)
    worksheet.merge_range('I1:P2', translate_text('הגשות', user, "hebrew"), title_format)
    worksheet.merge_range('Q1:X2', dates[0].strftime("%d.%m") + "-" + dates[-1].strftime("%d.%m"),
                          title_format)
    worksheet.write(2, 0, translate_text("תאריך", user, "hebrew"), cell_format)
    col = 1
    for d in dates:
        worksheet.write(2, col, d.strftime("%d.%m"), cell_format)
        col += 1
    worksheet.write(3, 0, translate_text("יום", user, "hebrew"), cell_format)
    col = 1
    for i in range(int(len(dates) / 7)):
        for d in days:
            worksheet.write(3, col, d, cell_format)
            col += 1
    worksheet.merge_range(f'A5:A{5 + maxes["morning"]}', translate_text('בוקר', user, "hebrew"), cell_format)
    worksheet.merge_range(
        f'A{5 + maxes["morning"] + 1}:A{5 + maxes["morning"] + maxes["after_noon"]}', translate_text('צהריים', user, "hebrew"), cell_format)
    worksheet.merge_range(
        f'A{5 + maxes["morning"] + maxes["after_noon"] + 1}:A{5 + maxes["morning"] + maxes["after_noon"] + maxes["night"]}',
        translate_text('לילה', user, "hebrew"), cell_format)
    start_extra = len(dates) + 3
    worksheet.merge_range(f'{get_column_letter(start_extra)}4:{get_column_letter(start_extra + 1)}4', 'שם', cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 2)}4", translate_text('בוקר', user, "hebrew") + " 1", cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 3)}4", translate_text('בוקר', user, "hebrew") + " 2", cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 4)}4", translate_text('צהריים', user, "hebrew") + " 1", cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 5)}4", translate_text('צהריים', user, "hebrew") + " 2", cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 6)}4", translate_text('לילה', user, "hebrew"), cell_format)
    worksheet.write(f"{get_column_letter(start_extra + 7)}4", translate_text("סופ\"ש", user, "hebrew"), cell_format)

    # Adding Data
    users = []
    row = 4
    col = 1
    for key in served:
        if key.count("M"):
            row = 4
            split = served[key].split("\n")
            for x in range(len(split)):
                if split[x] != "(לא משיכה)":
                    if x + 1 < len(split):
                        if split[x + 1] == "(לא משיכה)":
                            worksheet.write(row, col, split[x], cell_no_pull_format)
                        else:
                            worksheet.write(row, col, split[x])
                    else:
                        worksheet.write(row, col, split[x])
                    if split[x] not in users:
                        users.append(split[x])
                    row += 1
        elif key.count("A"):
            row = 4 + maxes["morning"] + 1
            split = served[key].split("\n")
            for x in range(len(split)):
                worksheet.write(row, col, split[x])
                if split[x] not in users:
                    users.append(split[x])
                row += 1
        else:
            row = 4 + maxes["morning"] + maxes["after_noon"] + 1
            split = served[key].split("\n")
            for x in range(len(split)):
                worksheet.write(row, col, split[x])
                if split[x] not in users:
                    users.append(split[x])
                row += 1
            col += 1

    num_rows = len(users) + 1
    for x in range(num_rows):
        col = len(dates) + 4
        if x == 0:
            worksheet.merge_range(f'{get_column_letter(col - 1)}{4 + x + 1}:{get_column_letter(col)}{4 + x + 1}', '', cell_format)
        else:
            worksheet.merge_range(f'{get_column_letter(col - 1)}{4 + x + 1}:{get_column_letter(col)}{4 + x + 1}', users[x - 1], cell_format)
        for c in range(6):
            worksheet.write(4 + x, col + c, "", cell_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{4 + num_rows + 1}:{get_column_letter(start_extra + 1)}{4 + num_rows + 1}',
                          translate_text('סה\"כ', user, "hebrew"), cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 2)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 2)}5:{get_column_letter(start_extra + 2)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 3)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 3)}5:{get_column_letter(start_extra + 3)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 4)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 4)}5:{get_column_letter(start_extra + 4)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 5)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 5)}5:{get_column_letter(start_extra + 5)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 6)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 6)}5:{get_column_letter(start_extra + 6)}{4 + num_rows})', cell_format)
    worksheet.write(f'{get_column_letter(start_extra + 7)}{4 + num_rows + 1}', f'=SUM({get_column_letter(start_extra + 7)}5:{get_column_letter(start_extra + 7)}{4 + num_rows})', cell_format)

    row = 4 + num_rows + 4

    worksheet.merge_range(f'{get_column_letter(start_extra)}{row}:{get_column_letter(start_extra + 7)}{row + 3}', translate_text('משמרות לאיכות', user, "hebrew"), title_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 4}:{get_column_letter(start_extra + 7)}{row + 5}', translate_text('שבוע ראשון', user, "hebrew"), title_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 6}:{get_column_letter(start_extra + 7)}{row + 6}', '', cell_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 7}:{get_column_letter(start_extra + 7)}{row + 7}', '', cell_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 8}:{get_column_letter(start_extra + 7)}{row + 9}', translate_text('שבוע שני', user, "hebrew"), title_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 10}:{get_column_letter(start_extra + 7)}{row + 10}', '', cell_format)
    worksheet.merge_range(f'{get_column_letter(start_extra)}{row + 11}:{get_column_letter(start_extra + 7)}{row + 11}', '', cell_format)

    row = row + 13
    count = 0
    worksheet.merge_range(
        f'{get_column_letter(start_extra)}{row + count}:{get_column_letter(start_extra + 7)}{row + count + 1}',
        translate_text('הערות', user, "hebrew"), title_format)
    count += 2
    split = notes_general.split("\n")
    if len(split) > 0:
        for s in split:
            worksheet.merge_range(
                f'{get_column_letter(start_extra)}{row + count}:{get_column_letter(start_extra + 7)}{row + count}', s,
                cell_format)
            count += 1
    for n in range(len(notes)):
        worksheet.merge_range(f'{get_column_letter(start_extra)}{row + count}:{get_column_letter(start_extra + 7)}{row + count}', translate_text(f'שבוע {str(n + 1)}', user, "hebrew"), title_format)
        count += 1
        split = notes[n].split("\n")
        if len(split) > 0:
            for s in split:
                worksheet.merge_range(f'{get_column_letter(start_extra)}{row + count}:{get_column_letter(start_extra + 7)}{row + count}', s, cell_format)
                count += 1

    worksheet.merge_range(f'{get_column_letter(start_extra + 10)}4:{get_column_letter(start_extra + 15)}5', translate_text('אירועים', user, "hebrew"), title_format)
    events = Event.objects.all()
    events_notes = []
    temp = ""
    for x in range(len(dates)):
        if len(events.filter(date2=dates[x])) > 0:
            for ev in events.filter(date2=dates[x]):
                if ev.nickname != "כולם":
                    events_notes.append(translate_text(f'בתאריך {ev.date2} יש {ev.description} ל{ev.nickname}', user, "hebrew"))
                else:
                    events_notes.append(translate_text(f'בתאריך {ev.date2} יש {ev.description}', user, "hebrew"))
    row = 6
    count = 0
    for s in events_notes:
        worksheet.merge_range(f'{get_column_letter(start_extra + 10)}{row + count}:{get_column_letter(start_extra + 15)}{row + count}', s, cell_format)
        count += 1

    workbook.close()
    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    file_name = "serve" + dates[0].strftime("%d.%m")
    return FileResponse(buffer, as_attachment=True, filename=f'{file_name}.xlsx')

# Compare different organization suggestions
def compare_organizations(served, guards_num, organization, officer, sat_night, users, users_settings):
    organizer = Organizer(served, guards_num, organization, officer, sat_night, users, users_settings)
    organizer.organize()
    for i in range(1):
        new_organizer = Organizer(served, guards_num, organization, officer, sat_night, users, users_settings)
        new_organizer.organize()
        if organizer.notes > new_organizer.notes:
            organizer = new_organizer
        if organizer.notes == 0:
            break
    return organizer



# template function filters


@register.filter
def check_user(shift, user):
    user_settings = USettings.objects.get(user=user)
    split = shift.replace("\r", "\n").split("\n")
    for s in split:
        s = s.replace(" ", "")
        if s == user_settings.nickname:
            return True
    return False

@register.filter
def get_served_week(served, week_num):
    return served[week_num]

@register.filter
def get_list_dict_served(served, shift):
    new_list = []
    for i in range(1, 8):
        new_list.append(served[f'{shift}{i}'])
    return new_list

@register.filter
def get_string_int(list, string):
    return list[int(string)]

@register.filter
def get_arming_name(log, request):
    return log.data[str(request.input_num)]["name"]


@register.filter
def arming_day_break(arming):
    keys = ["gun_id","user_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out"]
    input_num = arming.data["input_num"]
    new_data = []
    for i in range(1, input_num + 1):
        arming.data[str(i)]["id"] = i
        new_data.append(arming.data[str(i)])
    return new_data

@register.filter
def order_arming_dict_shift(arming):
    keys = ["gun_id","user_id", "name", "shift_num", "id_num", "time_in", "num_mags", "hand_cuffs", "gun_case", "mag_case", "keys", "radio", "radio_kit", "time_out"]  
    data = arming.data
    new_order = {}
    for i in range(int(data["input_num"])):
        new_order[i] = []
        for key in keys:
            new_order[i].append(data[f'{key}@{i}'])
    new_order = sorted(new_order.items(), key=lambda e: (e[1][3], e[1][5]))
    return new_order

@register.filter
def order_arming_dict_time(arming):
    pass


@register.filter
def check_keys(key, dictionary):
    for i in range(1, 8):
        if f'{i}@{key.id}' in dictionary.keys():
            return True
    return False

@register.filter
def get_week_shift(shift, week):
    shifts = []
    for i in range(1, 8):
        if f'{i}@{shift.id}' in week.keys():
            shifts.append(week[f'{i}@{shift.id}'])
        else:
            break
    return shifts

@register.filter
def get_city(string):
    settings = Settings.objects.all().first()
    return settings.city

@register.filter
def validation_log_check(log, shift):
    if log != None:
        if shift == 1:
            if log.name_checked_m != "" and log.name_checked_m != None:
                return True
        elif shift == 2:
            if log.name_checked_a != "" and log.name_checked_a != None:
                return True
        else:
            if log.name_checked_n != "" and log.name_checked_n != None:
                return True
    return False

@register.filter
def isStaff(user):
    return user.groups.filter(name='staff').exists()

@register.filter
def is_manager(user):
    return user.groups.filter(name='manager').exists()

@register.filter
def merge_user(obj_list, user):
    return [user, obj_list]

@register.filter
def object_list_cutter(data, reqtype):
    user = data[0]
    obj_list = data[1]
    if reqtype == "":
        new_list = []
        for day in obj_list:
            date = day.date
            day = arming_day_break(day)
            for row in day:
                if row["user_id"] == user.id:
                    row["date"] = date 
                    new_list.append(row)
    else:
        new_list = []
        for day in obj_list:
            date = day.date
            day = arming_day_break(day)
            for row in day:
                row["date"] = date
                new_list.append(row)
    return new_list

@register.filter
def countshifts(obj_list, user):
    count = 0
    for obj in obj_list:
        obj =  arming_day_break(obj)
        for log in obj:
            if log["user_id"] == user.id:
                count += 1
    return count

@register.filter
def gethours(obj_list, user):
    time_go = datetime.timedelta(0)
    for log in obj_list:
        log = arming_day_break(log)
        for row in log:
            if row["user_id"] == user.id:
                if row["time_out"] != None and row["time_out"] != "":
                    time_1 = datetime.datetime.strptime(row["time_in"],"%H:%M")
                    time_2 = datetime.datetime.strptime(row["time_out"],"%H:%M")
                    time_cal = time_2 - time_1
                    time_cal += datetime.timedelta(days=(time_cal.days * -1))
                    time_go += time_cal
    hours = time_go.days * 24
    temp_h = str(time_go)
    if time_go.days != 0:
        temp_h = temp_h.replace(f"{time_go.days} day, ", "")
    add_h = temp_h.split(":")
    hours += int(add_h[0])
    hours = str(hours) + "." + add_h[1]
    return hours


@register.filter
def getmonth(month):
    letter = month.lower()[0]
    if (letter >= 'a' and letter <= 'z'):
        return month
    else:
        months_he = ["ינו", "פבר", "מרץ", "אפר", "מאי", "יונ", "יול", "אוג", "ספט", "אוק", "נוב", "דצמ"]
        months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
        return months[months_he.index(month)]


@register.filter
def getfullname(user):
    return user.first_name + " " + user.last_name

@register.filter
def getday(string):
    letter = datetime.datetime.now().strftime("%b").lower()[0]
    if (letter >= 'a' and letter <= 'z'):
        return datetime.datetime.now()
    return datetime.datetime.now()

@register.filter
def is_users_log(user, log):
    if user.id == int(log["user_id"]):
        return True
    return False

@register.filter
def user_and_log(user, log):
    return [user, log]

@register.filter
def request_permission(arguments, log_obj):
    user = arguments[0]
    log = arguments[1]
    if user.id != int(log["user_id"]):
        return False
    armingrequests = ArmingRequest.objects.all().filter(log=log_obj, read=False, input_num=log["id"])
    if len(armingrequests) == 0:
        return True
    return False

@register.filter
def get_date_arming(user, day_obj):
    return [user, day_obj]

@register.filter
def edit_permission(arguments, log):
    user = arguments[0]
    log_obj = arguments[1]
    if user.groups.filter(name="manager").exists():
        return True
    if log["user_id"] == user.id:
        if int(log["shift_num"]) != 3:
            if datetime.datetime.now().date() == log_obj.date:
                return True
        else:
            if datetime.datetime.now().date() == log_obj.date or datetime.datetime.now().date() == log_obj.date + datetime.timedelta(days=1):
                return True
    return False

@register.filter
def timestr(time):
    if type(time) == Time:
        return time.strftime("%H:%M")
    else:
        return time

@register.filter
def num_to_shift(num):
    if num == 1:
        return "בוקר"
    elif num == 2:
        return "צהריים"
    return "לילה"

@register.filter
def counter_shifts(counter, arming_logs):
    morning = 0
    afternoon = 0
    night = 0
    arming_logs = arming_day_break(arming_logs)
    for log in arming_logs:
        if log["shift_num"] == 1:
            morning += 1
        elif log["shift_num"] == 2:
            afternoon += 1
        else:
            night += 1
    if counter <= morning:
        return counter
    elif counter <= morning + afternoon:
        return counter - morning
    else:
        return counter - morning - afternoon

@register.filter(name="translate_text")
def translate_text(text, user, from_language="hebrew"):
    if user.is_authenticated:
        user_settings = USettings.objects.all().filter(user=user).first()
        if from_language != user_settings.language:
            langs_dict = GoogleTranslator.get_supported_languages(as_dict=True)
            translator = GoogleTranslator(source='auto', target=langs_dict[user_settings.language])
            return translator.translate(text).capitalize()
    return text


@register.filter
def translate_text_batch(texts, user, from_language):
    translated = []
    if user.is_authenticated:
        user_settings = USettings.objects.all().filter(user=user).first()
        if from_language != user_settings.language:
            langs_dict = GoogleTranslator.get_supported_languages(as_dict=True)
            translator = GoogleTranslator(source='auto', target=langs_dict[user_settings.language])
            for text in texts:
                translated.append(translator.translate(text).capitalize())
            return translated
    return texts


@register.filter
def get_base_string(user, num):
    return translate_text(base_strings[int(num)], user, "hebrew")


@register.filter
def get_user_name(username):
    user = User.objects.all().filter(username=username).first()
    name = user.first_name + " " + user.last_name
    return name


@register.filter
def get_item(dictionary, key):
    return dictionary.get(key)


@register.filter
def is_divided5(num):
    if num % 5 == 0:
        return True
    return False


@register.filter
def get_index(list, item):
    return list.index(item)


@register.filter
def minus(item, num):
    return item - num


@register.filter
def plus_days(date):
    return date + datetime.timedelta(days=13)


@register.filter
def cut_list(list, num):
    new_list = []
    if num == 1:
        for x in range(7):
            new_list.append(list[x])
    else:
        for x in range(7, len(list)):
            new_list.append(list[x])
    return new_list


@register.filter
def is_in_td(text, nickname):
    if text is not None and len(text) > 0:
        split = text.split("\n")
        for s in split:
            s = s.replace(" ", "")
            s = s.replace("\n", "")
            s = s.replace("\r", "")
            if nickname == s:
                return True
    return False


@register.filter
def is_string(item):
    return isinstance(item, str)


@register.filter
def clip_dictionary(served, num_week):
    days = range((num_week * 7) + 1, (num_week * 7) + 8)
    new_list = []
    for day in days:
        new_list.append(served[f'day{day}'])
    return new_list


@register.filter
def clip_dictionary_served1(served, kind):
    new_list = {}
    for key in served.keys():
        if kind in key:
            new_list[key] = served[key]
    return new_list


@register.filter
def clip_dictionary_served2(served, num_week):
    kind = "M"
    for key in served.keys():
        if kind in key:
            break
        elif 'A' in key:
            kind = 'A'
            break
        elif 'N' in key:
            kind = 'N'
            break
    days = range((num_week * 7) + 1, (num_week * 7) + 8)
    new_list = []
    for day in days:
        new_list.append(served[kind + str(day)])
    return new_list


@register.filter
def clip_days(days, num_week):
    new_days = []
    for i in range(num_week * 7, num_week * 7 + 7):
        new_days.append(days[i])
    return new_days


@register.filter
def to_array(end, start):
    start = int(start)
    end = int(end)
    return range(start, end)


@register.filter
def to_array2(start, end):
    start = int(start)
    end = int(end)
    return range(start, end)


@register.filter
def get_form_data(form, kind):
    array = []
    for i in range(1, 8):
        array.append(getattr(form.instance, f"{kind}{i}"))
    return array


@register.filter
def get_days(organization):
    days = []
    for x in range(organization.num_weeks * 7):
        days.append(organization.date + datetime.timedelta(days=x))
    return days


@register.filter
def get_weeks(organization):
    weeks = []
    for i in range(organization.num_weeks):
        weeks.append(organization.weeks_data[str(i)])
    return weeks


@register.filter
def get_num_organization(organization):
    organizations = Organization.objects.all().order_by('-date')
    i = 0
    for org in organizations:
        if org.date == organization.date:
            return i
        i += 1
    return -1
